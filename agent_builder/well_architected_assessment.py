#!/usr/bin/env python3
"""
GitHub Well-Architected Framework — Organizational Assessment Script

A comprehensive, interactive assessment that evaluates a GitHub organization's
repositories against the five pillars of the Well-Architected Framework:
  1. Security
  2. Reliability
  3. Operational Excellence
  4. Performance Efficiency
  5. Collaboration & Developer Experience

Usage:
    python well_architected_assessment.py [--repo-path /path/to/repo] [--output report.md]

The script combines:
  - Automated file/config scanning of the local repository
  - Interactive organizational-policy questions (for things that can't be detected from files)
  - GitHub API checks (optional, requires GITHUB_TOKEN)
  - Detailed scoring, gap analysis, and remediation guidance
"""

from __future__ import annotations

import argparse
import glob
import html as html_mod
import json
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Optional

# ---------------------------------------------------------------------------
# Constants & Types
# ---------------------------------------------------------------------------

FRAMEWORK_URL = "https://wellarchitected.github.com"

class Status(Enum):
    PASS = "PASS"
    PARTIAL = "PARTIAL"
    FAIL = "FAIL"
    ACCEPTABLE_RISK = "ACCEPTABLE RISK"
    ALTERNATIVE = "ALTERNATIVE APPROACH"
    NOT_APPLICABLE = "N/A"

class Severity(Enum):
    CRITICAL = "Critical"
    HIGH = "High"
    MEDIUM = "Medium"
    LOW = "Low"
    INFO = "Info"

@dataclass
class Finding:
    pillar: str
    principle: str
    status: Status
    severity: Severity
    evidence: str
    recommendation: str
    acceptable_risk_note: str = ""
    alternative_note: str = ""

@dataclass
class PillarScore:
    name: str
    findings: list[Finding] = field(default_factory=list)

    @property
    def score(self) -> float:
        if not self.findings:
            return 0.0
        applicable = [f for f in self.findings if f.status != Status.NOT_APPLICABLE]
        if not applicable:
            return 10.0
        points = 0.0
        for f in applicable:
            if f.status == Status.PASS:
                points += 1.0
            elif f.status in (Status.PARTIAL, Status.ACCEPTABLE_RISK, Status.ALTERNATIVE):
                points += 0.5
        return round((points / len(applicable)) * 10, 1)

    @property
    def compliance(self) -> str:
        s = self.score
        if s >= 8:
            return "COMPLIANT"
        elif s >= 5:
            return "PARTIAL"
        return "NON-COMPLIANT"


# ---------------------------------------------------------------------------
# Terminal helpers
# ---------------------------------------------------------------------------

COLORS = {
    "green": "\033[92m",
    "yellow": "\033[93m",
    "red": "\033[91m",
    "cyan": "\033[96m",
    "bold": "\033[1m",
    "dim": "\033[2m",
    "reset": "\033[0m",
}

def _c(text: str, color: str) -> str:
    if not sys.stdout.isatty():
        return text
    return f"{COLORS.get(color, '')}{text}{COLORS['reset']}"

def banner(text: str) -> None:
    width = 72
    print()
    print(_c("=" * width, "cyan"))
    print(_c(f"  {text}", "bold"))
    print(_c("=" * width, "cyan"))
    print()

def section(text: str) -> None:
    print()
    print(_c(f"── {text} ──", "cyan"))
    print()

def info(text: str) -> None:
    print(f"  {_c('ℹ', 'cyan')} {text}")

def success(text: str) -> None:
    print(f"  {_c('✓', 'green')} {text}")

def warn(text: str) -> None:
    print(f"  {_c('⚠', 'yellow')} {text}")

def fail(text: str) -> None:
    print(f"  {_c('✗', 'red')} {text}")

def status_icon(status: Status) -> str:
    icons = {
        Status.PASS: _c("✓", "green"),
        Status.PARTIAL: _c("◐", "yellow"),
        Status.FAIL: _c("✗", "red"),
        Status.ACCEPTABLE_RISK: _c("⚡", "yellow"),
        Status.ALTERNATIVE: _c("↔", "cyan"),
        Status.NOT_APPLICABLE: _c("—", "dim"),
    }
    return icons.get(status, "?")


# ---------------------------------------------------------------------------
# Interactive question helpers
# ---------------------------------------------------------------------------

def ask_yes_no(question: str, default: Optional[bool] = None) -> bool:
    """Ask a yes/no question and return the boolean result."""
    hint = "[Y/n]" if default is True else ("[y/N]" if default is False else "[y/n]")
    while True:
        answer = input(f"  {_c('?', 'cyan')} {question} {_c(hint, 'dim')}: ").strip().lower()
        if answer in ("y", "yes"):
            return True
        if answer in ("n", "no"):
            return False
        if answer == "" and default is not None:
            return default
        print(f"    Please answer 'y' or 'n'.")

def ask_choice(question: str, options: list[str], default: int = 0) -> int:
    """Ask a multiple-choice question and return the selected index."""
    print(f"\n  {_c('?', 'cyan')} {question}")
    for i, opt in enumerate(options):
        marker = _c("→", "green") if i == default else " "
        print(f"    {marker} {i + 1}. {opt}")
    while True:
        raw = input(f"    {_c('Choose', 'dim')} [1-{len(options)}, default={default + 1}]: ").strip()
        if raw == "":
            return default
        if raw.isdigit() and 1 <= int(raw) <= len(options):
            return int(raw) - 1
        print(f"    Please enter a number between 1 and {len(options)}.")

def ask_text(question: str, default: str = "") -> str:
    """Ask for free-text input."""
    hint = f" {_c(f'[{default}]', 'dim')}" if default else ""
    answer = input(f"  {_c('?', 'cyan')} {question}{hint}: ").strip()
    return answer if answer else default


# ---------------------------------------------------------------------------
# File-system scanning helpers
# ---------------------------------------------------------------------------

def find_file(repo: Path, *candidates: str) -> Optional[Path]:
    """Return the first candidate path that exists under `repo`."""
    for c in candidates:
        p = repo / c
        if p.exists():
            return p
    return None

def find_files(repo: Path, pattern: str) -> list[Path]:
    """Return all files matching a glob under `repo`."""
    return sorted(repo.glob(pattern))

def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")

def yaml_like_search(text: str, key: str) -> Optional[str]:
    """Cheap YAML key lookup (avoids requiring PyYAML)."""
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith(f"{key}:"):
            return stripped[len(f"{key}:"):].strip().strip('"').strip("'")
    return None

def file_contains(path: Path, pattern: str, flags: int = re.IGNORECASE) -> bool:
    return bool(re.search(pattern, read_text(path), flags))


# ---------------------------------------------------------------------------
# Automated checks
# ---------------------------------------------------------------------------

class RepoScanner:
    """Scans a local repository checkout for well-architected indicators."""

    def __init__(self, repo_path: Path):
        self.repo = repo_path
        self.github_dir = repo_path / ".github"
        self.workflows_dir = self.github_dir / "workflows"

    # -- Utility --

    def _workflow_files(self) -> list[Path]:
        return find_files(self.repo, ".github/workflows/*.yml") + \
               find_files(self.repo, ".github/workflows/*.yaml")

    def _all_workflow_text(self) -> str:
        return "\n".join(read_text(w) for w in self._workflow_files())

    # ==================== SECURITY ====================

    def check_dependabot_config(self) -> Finding:
        """Check for Dependabot configuration."""
        path = find_file(self.repo, ".github/dependabot.yml", ".github/dependabot.yaml")
        if path:
            text = read_text(path)
            ecosystems = re.findall(r'package-ecosystem:\s*["\']?(\S+)', text)
            if ecosystems:
                return Finding(
                    pillar="Security",
                    principle="Automated dependency updates (Dependabot)",
                    status=Status.PASS,
                    severity=Severity.HIGH,
                    evidence=f"dependabot.yml found with ecosystems: {', '.join(ecosystems)}",
                    recommendation="None — Dependabot is properly configured.",
                )
            return Finding(
                pillar="Security",
                principle="Automated dependency updates (Dependabot)",
                status=Status.PARTIAL,
                severity=Severity.HIGH,
                evidence="dependabot.yml exists but no package-ecosystem entries detected.",
                recommendation="Add at least one package-ecosystem entry for your project's language.",
            )
        return Finding(
            pillar="Security",
            principle="Automated dependency updates (Dependabot)",
            status=Status.FAIL,
            severity=Severity.HIGH,
            evidence="No .github/dependabot.yml found.",
            recommendation=(
                "Create .github/dependabot.yml to enable automated dependency updates. "
                "See: https://docs.github.com/en/code-security/dependabot/dependabot-version-updates"
            ),
            alternative_note=(
                "Renovate Bot is an acceptable alternative to Dependabot. "
                "If your org uses Renovate, this is an acceptable alternative approach."
            ),
        )

    def check_codeql_scanning(self) -> Finding:
        """Check for CodeQL or code scanning workflows."""
        workflows = self._workflow_files()
        for wf in workflows:
            text = read_text(wf)
            if re.search(r'github/codeql-action|codeql-action/init|codeql-action/analyze', text):
                return Finding(
                    pillar="Security",
                    principle="Code scanning with CodeQL",
                    status=Status.PASS,
                    severity=Severity.CRITICAL,
                    evidence=f"CodeQL scanning found in {wf.relative_to(self.repo)}",
                    recommendation="None — CodeQL is configured.",
                )
        # Check for alternative SAST tools
        all_wf = self._all_workflow_text()
        alt_tools = []
        for tool_name, pattern in [
            ("Semgrep", r"semgrep"),
            ("SonarQube/SonarCloud", r"sonar(?:qube|cloud|scanner)"),
            ("Snyk Code", r"snyk.*code|snyk/actions"),
            ("Checkmarx", r"checkmarx"),
        ]:
            if re.search(pattern, all_wf, re.IGNORECASE):
                alt_tools.append(tool_name)
        if alt_tools:
            return Finding(
                pillar="Security",
                principle="Code scanning with CodeQL",
                status=Status.ALTERNATIVE,
                severity=Severity.CRITICAL,
                evidence=f"Alternative SAST tool(s) detected: {', '.join(alt_tools)}",
                recommendation="CodeQL is recommended for GitHub-native integration, but your current SAST tool provides coverage.",
                alternative_note=f"Using {', '.join(alt_tools)} instead of CodeQL. This is acceptable if the tool covers your languages and is actively maintained.",
            )
        return Finding(
            pillar="Security",
            principle="Code scanning with CodeQL",
            status=Status.FAIL,
            severity=Severity.CRITICAL,
            evidence="No CodeQL or alternative SAST scanning workflow found.",
            recommendation=(
                "Enable CodeQL code scanning. Go to Settings → Code security → Code scanning, "
                "or add the github/codeql-action to a workflow. "
                "See: https://docs.github.com/en/code-security/code-scanning"
            ),
        )

    def check_security_md(self) -> Finding:
        """Check for SECURITY.md vulnerability disclosure policy."""
        path = find_file(self.repo, "SECURITY.md", ".github/SECURITY.md", "docs/SECURITY.md")
        if path:
            text = read_text(path)
            if len(text.strip()) > 50:
                return Finding(
                    pillar="Security",
                    principle="Vulnerability disclosure policy (SECURITY.md)",
                    status=Status.PASS,
                    severity=Severity.HIGH,
                    evidence=f"SECURITY.md found at {path.relative_to(self.repo)} with substantive content.",
                    recommendation="None — security policy is in place.",
                )
            return Finding(
                pillar="Security",
                principle="Vulnerability disclosure policy (SECURITY.md)",
                status=Status.PARTIAL,
                severity=Severity.HIGH,
                evidence="SECURITY.md exists but appears to be a stub.",
                recommendation="Flesh out SECURITY.md with reporting instructions, response SLA, and scope.",
            )
        return Finding(
            pillar="Security",
            principle="Vulnerability disclosure policy (SECURITY.md)",
            status=Status.FAIL,
            severity=Severity.HIGH,
            evidence="No SECURITY.md found.",
            recommendation=(
                "Create a SECURITY.md in the repository root with vulnerability reporting instructions. "
                "See: https://docs.github.com/en/code-security/getting-started/adding-a-security-policy-to-your-repository"
            ),
        )

    def check_token_permissions(self) -> Finding:
        """Check that workflows scope GITHUB_TOKEN permissions per-job."""
        workflows = self._workflow_files()
        if not workflows:
            return Finding(
                pillar="Security",
                principle="Least-privilege GITHUB_TOKEN permissions",
                status=Status.NOT_APPLICABLE,
                severity=Severity.HIGH,
                evidence="No workflow files found.",
                recommendation="Will be assessed when workflows are added.",
            )
        scoped_count = 0
        unscoped = []
        for wf in workflows:
            text = read_text(wf)
            if re.search(r'^\s*permissions:', text, re.MULTILINE):
                scoped_count += 1
            else:
                unscoped.append(str(wf.relative_to(self.repo)))
        if not unscoped:
            return Finding(
                pillar="Security",
                principle="Least-privilege GITHUB_TOKEN permissions",
                status=Status.PASS,
                severity=Severity.HIGH,
                evidence=f"All {scoped_count} workflow(s) define explicit permissions.",
                recommendation="None — permissions are properly scoped.",
            )
        if scoped_count > 0:
            return Finding(
                pillar="Security",
                principle="Least-privilege GITHUB_TOKEN permissions",
                status=Status.PARTIAL,
                severity=Severity.HIGH,
                evidence=f"{scoped_count}/{len(workflows)} workflows define permissions. Missing: {', '.join(unscoped[:5])}",
                recommendation="Add a top-level or per-job `permissions:` block to all workflows to enforce least-privilege.",
            )
        return Finding(
            pillar="Security",
            principle="Least-privilege GITHUB_TOKEN permissions",
            status=Status.FAIL,
            severity=Severity.HIGH,
            evidence="No workflows define explicit GITHUB_TOKEN permissions.",
            recommendation=(
                "Add `permissions:` blocks to every workflow. Start with `permissions: {}` (no access) "
                "and grant only what each job needs. See: https://docs.github.com/en/actions/security-for-github-actions"
            ),
        )

    def check_action_pinning(self) -> Finding:
        """Check that third-party actions are pinned to full SHAs (not tags)."""
        workflows = self._workflow_files()
        if not workflows:
            return Finding(
                pillar="Security",
                principle="Supply chain integrity — action pinning",
                status=Status.NOT_APPLICABLE,
                severity=Severity.HIGH,
                evidence="No workflow files found.",
                recommendation="Will be assessed when workflows are added.",
            )
        uses_pattern = re.compile(r'uses:\s*([^\s#]+)')
        total_uses = 0
        pinned = 0
        unpinned_examples = []
        for wf in workflows:
            for match in uses_pattern.finditer(read_text(wf)):
                ref = match.group(1)
                # Skip local actions (./), reusable workflows within org
                if ref.startswith("./") or ref.startswith(".github/"):
                    continue
                total_uses += 1
                # Check if pinned to SHA (40-char hex after @)
                if re.search(r'@[0-9a-f]{40}', ref):
                    pinned += 1
                elif len(unpinned_examples) < 5:
                    unpinned_examples.append(ref)
        if total_uses == 0:
            return Finding(
                pillar="Security",
                principle="Supply chain integrity — action pinning",
                status=Status.NOT_APPLICABLE,
                severity=Severity.HIGH,
                evidence="No third-party action references found.",
                recommendation="N/A.",
            )
        if pinned == total_uses:
            return Finding(
                pillar="Security",
                principle="Supply chain integrity — action pinning",
                status=Status.PASS,
                severity=Severity.HIGH,
                evidence=f"All {total_uses} action references are pinned to full SHAs.",
                recommendation="None — excellent supply chain hygiene.",
            )
        ratio = pinned / total_uses
        return Finding(
            pillar="Security",
            principle="Supply chain integrity — action pinning",
            status=Status.PARTIAL if ratio > 0.5 else Status.FAIL,
            severity=Severity.HIGH,
            evidence=f"{pinned}/{total_uses} actions pinned to SHA. Unpinned: {', '.join(unpinned_examples)}",
            recommendation=(
                "Pin all third-party actions to their full commit SHA instead of a tag. "
                "Use `actions/checkout@<full-sha>` with a comment noting the version. "
                "Tools like StepSecurity/secure-repo can automate this."
            ),
            acceptable_risk_note=(
                "Pinning to major version tags (e.g., @v4) is a common acceptable risk for "
                "first-party GitHub actions (actions/*) since GitHub controls them. "
                "Third-party actions should always be SHA-pinned."
            ),
        )

    def check_oidc_usage(self) -> Finding:
        """Check for OIDC-based cloud auth instead of long-lived secrets."""
        all_wf = self._all_workflow_text()
        if not all_wf:
            return Finding(
                pillar="Security",
                principle="OIDC for cloud deployments",
                status=Status.NOT_APPLICABLE,
                severity=Severity.MEDIUM,
                evidence="No workflows found to assess.",
                recommendation="Will be assessed when cloud deployment workflows are added.",
            )
        # Detect cloud deployment patterns
        has_aws = bool(re.search(r'aws-actions/configure-aws-credentials|AWS_ACCESS_KEY|aws-region', all_wf, re.IGNORECASE))
        has_azure = bool(re.search(r'azure/login|AZURE_CREDENTIALS|azure/webapps-deploy', all_wf, re.IGNORECASE))
        has_gcp = bool(re.search(r'google-github-actions/auth|GOOGLE_CREDENTIALS|gcloud', all_wf, re.IGNORECASE))

        if not (has_aws or has_azure or has_gcp):
            return Finding(
                pillar="Security",
                principle="OIDC for cloud deployments",
                status=Status.NOT_APPLICABLE,
                severity=Severity.MEDIUM,
                evidence="No cloud deployment patterns detected in workflows.",
                recommendation="N/A — no cloud integrations found.",
            )

        uses_oidc = bool(re.search(
            r'id-token:\s*write|role-to-assume|audience:|workload_identity_provider',
            all_wf, re.IGNORECASE
        ))
        uses_long_lived = bool(re.search(
            r'AWS_ACCESS_KEY_ID|AWS_SECRET_ACCESS_KEY|AZURE_CREDENTIALS.*\$\{\{|GOOGLE_CREDENTIALS.*\$\{\{',
            all_wf
        ))

        if uses_oidc and not uses_long_lived:
            return Finding(
                pillar="Security",
                principle="OIDC for cloud deployments",
                status=Status.PASS,
                severity=Severity.MEDIUM,
                evidence="OIDC-based authentication detected; no long-lived secret patterns found.",
                recommendation="None — OIDC is properly configured.",
            )
        if uses_oidc and uses_long_lived:
            return Finding(
                pillar="Security",
                principle="OIDC for cloud deployments",
                status=Status.PARTIAL,
                severity=Severity.MEDIUM,
                evidence="OIDC detected alongside long-lived secret references. Migration may be in progress.",
                recommendation="Complete the migration from long-lived secrets to OIDC for all cloud integrations.",
            )
        return Finding(
            pillar="Security",
            principle="OIDC for cloud deployments",
            status=Status.FAIL,
            severity=Severity.MEDIUM,
            evidence="Cloud deployment detected using long-lived secrets instead of OIDC.",
            recommendation=(
                "Switch to OIDC (OpenID Connect) for cloud authentication. "
                "AWS: use `role-to-assume` with `aws-actions/configure-aws-credentials`. "
                "Azure: use `azure/login` with federated credentials. "
                "GCP: use `google-github-actions/auth` with workload identity."
            ),
            acceptable_risk_note=(
                "Some environments (self-hosted runners, on-prem deployments) may not support OIDC. "
                "In those cases, ensure secrets are rotated regularly and scoped to the minimum required permissions."
            ),
        )

    # ==================== RELIABILITY ====================

    def check_ci_workflows(self) -> Finding:
        """Check for CI/testing workflows."""
        workflows = self._workflow_files()
        ci_indicators = []
        for wf in workflows:
            text = read_text(wf)
            if re.search(r'pull_request|push.*main|push.*master', text) and \
               re.search(r'test|lint|check|build|ci', text, re.IGNORECASE):
                ci_indicators.append(str(wf.relative_to(self.repo)))
        if ci_indicators:
            return Finding(
                pillar="Reliability",
                principle="Automated CI testing before merge",
                status=Status.PASS,
                severity=Severity.CRITICAL,
                evidence=f"CI workflow(s) found: {', '.join(ci_indicators[:3])}",
                recommendation="None — CI is in place.",
            )
        if workflows:
            return Finding(
                pillar="Reliability",
                principle="Automated CI testing before merge",
                status=Status.PARTIAL,
                severity=Severity.CRITICAL,
                evidence="Workflows exist but none appear to run tests on pull requests.",
                recommendation="Add a workflow triggered on `pull_request` that runs your test suite.",
            )
        return Finding(
            pillar="Reliability",
            principle="Automated CI testing before merge",
            status=Status.FAIL,
            severity=Severity.CRITICAL,
            evidence="No CI workflows found.",
            recommendation=(
                "Create a CI workflow that runs on `pull_request` and `push` to the default branch. "
                "Include linting, unit tests, and build verification."
            ),
            alternative_note=(
                "If CI runs in an external system (Jenkins, CircleCI, etc.), ensure results are reported "
                "back to GitHub via commit statuses or check runs so branch protection can enforce them."
            ),
        )

    def check_rollback_strategy(self) -> Finding:
        """Check for rollback or revert workflows."""
        all_wf = self._all_workflow_text()
        has_rollback = bool(re.search(
            r'rollback|revert|undo|previous.version|canary|blue.green|deployment.*slot',
            all_wf, re.IGNORECASE
        ))
        if has_rollback:
            return Finding(
                pillar="Reliability",
                principle="Rollback / revert strategy",
                status=Status.PASS,
                severity=Severity.HIGH,
                evidence="Rollback/revert patterns detected in workflows.",
                recommendation="None — rollback strategy is present.",
            )
        # Check for deployment workflows that might need rollback
        has_deploy = bool(re.search(r'deploy|release|publish', all_wf, re.IGNORECASE))
        if has_deploy:
            return Finding(
                pillar="Reliability",
                principle="Rollback / revert strategy",
                status=Status.FAIL,
                severity=Severity.HIGH,
                evidence="Deployment workflows found but no rollback or revert mechanism detected.",
                recommendation=(
                    "Add a rollback workflow or strategy. Options: "
                    "(1) A manual `workflow_dispatch` rollback workflow, "
                    "(2) Blue-green / canary deployments, "
                    "(3) Helm rollback or platform-native rollback commands."
                ),
                acceptable_risk_note=(
                    "If your deployment platform (e.g., Kubernetes, Azure App Service) has built-in "
                    "rollback capabilities, you may rely on those instead of a GitHub workflow. "
                    "Document this decision in your runbook."
                ),
            )
        return Finding(
            pillar="Reliability",
            principle="Rollback / revert strategy",
            status=Status.NOT_APPLICABLE,
            severity=Severity.HIGH,
            evidence="No deployment workflows detected.",
            recommendation="Will be assessed when deployment workflows are added.",
        )

    def check_environments(self) -> Finding:
        """Check for environment references in workflows (deployment protection)."""
        all_wf = self._all_workflow_text()
        env_refs = re.findall(r'environment:\s*(\S+)', all_wf)
        if env_refs:
            unique_envs = list(set(e.strip("'\"") for e in env_refs))
            has_prod_protection = any(
                e.lower() in ("production", "prod", "staging", "stage")
                for e in unique_envs
            )
            if has_prod_protection:
                return Finding(
                    pillar="Reliability",
                    principle="Deployment environments with protection rules",
                    status=Status.PASS,
                    severity=Severity.HIGH,
                    evidence=f"Environment references found: {', '.join(unique_envs)}",
                    recommendation="Verify that production/staging environments have required reviewers and wait timers in GitHub Settings → Environments.",
                )
            return Finding(
                pillar="Reliability",
                principle="Deployment environments with protection rules",
                status=Status.PARTIAL,
                severity=Severity.HIGH,
                evidence=f"Environments referenced ({', '.join(unique_envs)}) but no obvious production environment with protection.",
                recommendation="Define a 'production' environment in Settings → Environments with required reviewers.",
            )
        deploy_detected = bool(re.search(r'deploy|release.*publish', all_wf, re.IGNORECASE))
        if deploy_detected:
            return Finding(
                pillar="Reliability",
                principle="Deployment environments with protection rules",
                status=Status.FAIL,
                severity=Severity.HIGH,
                evidence="Deployment workflows exist but no GitHub Environments are referenced.",
                recommendation=(
                    "Use GitHub Environments to add protection rules (required reviewers, wait timers) "
                    "for production deployments. See: https://docs.github.com/en/actions/deployment/targeting-different-environments"
                ),
            )
        return Finding(
            pillar="Reliability",
            principle="Deployment environments with protection rules",
            status=Status.NOT_APPLICABLE,
            severity=Severity.HIGH,
            evidence="No deployment patterns detected.",
            recommendation="N/A.",
        )

    # ==================== OPERATIONAL EXCELLENCE ====================

    def check_contributing_md(self) -> Finding:
        """Check for CONTRIBUTING.md."""
        path = find_file(self.repo, "CONTRIBUTING.md", ".github/CONTRIBUTING.md", "docs/CONTRIBUTING.md")
        if path:
            text = read_text(path)
            if len(text.strip()) > 100:
                return Finding(
                    pillar="Operational Excellence",
                    principle="Contributing guidelines (CONTRIBUTING.md)",
                    status=Status.PASS,
                    severity=Severity.MEDIUM,
                    evidence=f"CONTRIBUTING.md found at {path.relative_to(self.repo)} ({len(text)} chars).",
                    recommendation="None — contributing guide is present.",
                )
            return Finding(
                pillar="Operational Excellence",
                principle="Contributing guidelines (CONTRIBUTING.md)",
                status=Status.PARTIAL,
                severity=Severity.MEDIUM,
                evidence="CONTRIBUTING.md exists but appears to be a stub.",
                recommendation="Expand with setup instructions, coding standards, PR process, and testing requirements.",
            )
        return Finding(
            pillar="Operational Excellence",
            principle="Contributing guidelines (CONTRIBUTING.md)",
            status=Status.FAIL,
            severity=Severity.MEDIUM,
            evidence="No CONTRIBUTING.md found.",
            recommendation="Create a CONTRIBUTING.md covering: setup, coding standards, branching strategy, PR process, and testing.",
        )

    def check_pr_template(self) -> Finding:
        """Check for pull request templates."""
        candidates = [
            ".github/PULL_REQUEST_TEMPLATE.md",
            ".github/pull_request_template.md",
            "PULL_REQUEST_TEMPLATE.md",
            "docs/pull_request_template.md",
        ]
        path = find_file(self.repo, *candidates)
        # Also check for a template directory
        template_dir = self.repo / ".github" / "PULL_REQUEST_TEMPLATE"
        has_template_dir = template_dir.is_dir() and any(template_dir.iterdir())

        if path or has_template_dir:
            return Finding(
                pillar="Operational Excellence",
                principle="Pull request templates",
                status=Status.PASS,
                severity=Severity.MEDIUM,
                evidence=f"PR template found: {path.relative_to(self.repo) if path else 'template directory'}",
                recommendation="None — PR template is in place.",
            )
        return Finding(
            pillar="Operational Excellence",
            principle="Pull request templates",
            status=Status.FAIL,
            severity=Severity.MEDIUM,
            evidence="No pull request template found.",
            recommendation=(
                "Create .github/PULL_REQUEST_TEMPLATE.md with sections for: "
                "description, type of change, testing, checklist."
            ),
        )

    def check_reusable_workflows(self) -> Finding:
        """Check for reusable workflow patterns."""
        workflows = self._workflow_files()
        if not workflows:
            return Finding(
                pillar="Operational Excellence",
                principle="Reusable workflows to reduce duplication",
                status=Status.NOT_APPLICABLE,
                severity=Severity.MEDIUM,
                evidence="No workflows found.",
                recommendation="N/A.",
            )
        reusable_count = 0
        caller_count = 0
        for wf in workflows:
            text = read_text(wf)
            if re.search(r'workflow_call', text):
                reusable_count += 1
            if re.search(r'uses:\s*\./.github/workflows/|uses:\s*\w+/\w+/\.github/workflows/', text):
                caller_count += 1
        if reusable_count > 0 or caller_count > 0:
            return Finding(
                pillar="Operational Excellence",
                principle="Reusable workflows to reduce duplication",
                status=Status.PASS,
                severity=Severity.MEDIUM,
                evidence=f"{reusable_count} reusable workflow(s), {caller_count} caller(s) detected.",
                recommendation="None — reusable workflow patterns are in use.",
            )
        if len(workflows) >= 3:
            return Finding(
                pillar="Operational Excellence",
                principle="Reusable workflows to reduce duplication",
                status=Status.PARTIAL,
                severity=Severity.MEDIUM,
                evidence=f"{len(workflows)} workflows found but no reusable workflow pattern detected.",
                recommendation="Consider extracting shared steps into reusable workflows (workflow_call) to reduce duplication.",
                acceptable_risk_note=(
                    "For small repos with few workflows, the overhead of reusable workflows may not be justified. "
                    "This becomes more important as the number of repos and workflows grows."
                ),
            )
        return Finding(
            pillar="Operational Excellence",
            principle="Reusable workflows to reduce duplication",
            status=Status.PARTIAL,
            severity=Severity.LOW,
            evidence=f"Only {len(workflows)} workflow(s) — limited opportunity for reuse.",
            recommendation="Consider reusable workflows as the project grows.",
            acceptable_risk_note="Few workflows — limited benefit from reusable workflow patterns at this scale.",
        )

    def check_readme_quality(self) -> Finding:
        """Check README.md quality and completeness."""
        path = find_file(self.repo, "README.md", "readme.md")
        if not path:
            return Finding(
                pillar="Operational Excellence",
                principle="Quality documentation (README.md)",
                status=Status.FAIL,
                severity=Severity.HIGH,
                evidence="No README.md found.",
                recommendation="Create a comprehensive README.md with project description, setup instructions, usage, and contribution guidelines.",
            )
        text = read_text(path)
        score_items = {
            "badges/shields": bool(re.search(r'!\[.*\]\(.*shields\.io|badge|img\.shields', text)),
            "installation/setup section": bool(re.search(r'##.*(?:install|setup|getting.started|prerequisites)', text, re.IGNORECASE)),
            "usage section": bool(re.search(r'##.*(?:usage|how.to|quick.start|example)', text, re.IGNORECASE)),
            "license reference": bool(re.search(r'license', text, re.IGNORECASE)),
            "substantive content (>500 chars)": len(text.strip()) > 500,
        }
        found = [k for k, v in score_items.items() if v]
        missing = [k for k, v in score_items.items() if not v]
        if len(found) >= 4:
            return Finding(
                pillar="Operational Excellence",
                principle="Quality documentation (README.md)",
                status=Status.PASS,
                severity=Severity.HIGH,
                evidence=f"README.md is comprehensive. Has: {', '.join(found)}.",
                recommendation="None — README is solid.",
            )
        if len(found) >= 2:
            return Finding(
                pillar="Operational Excellence",
                principle="Quality documentation (README.md)",
                status=Status.PARTIAL,
                severity=Severity.HIGH,
                evidence=f"README.md present but could be improved. Missing: {', '.join(missing)}.",
                recommendation=f"Add sections for: {', '.join(missing)}.",
            )
        return Finding(
            pillar="Operational Excellence",
            principle="Quality documentation (README.md)",
            status=Status.PARTIAL,
            severity=Severity.HIGH,
            evidence=f"README.md is minimal ({len(text)} chars). Missing: {', '.join(missing)}.",
            recommendation="Expand README with installation, usage, contributing, and license sections.",
        )

    def check_stale_automation(self) -> Finding:
        """Check for stale issue/PR automation."""
        all_wf = self._all_workflow_text()
        has_stale = bool(re.search(r'actions/stale|stale-action|close-stale|label.*stale', all_wf, re.IGNORECASE))
        # Also check for GitHub app-based stale management
        probot_config = find_file(self.repo, ".github/stale.yml")
        if has_stale or probot_config:
            return Finding(
                pillar="Operational Excellence",
                principle="Stale issue/PR automation",
                status=Status.PASS,
                severity=Severity.LOW,
                evidence="Stale issue/PR automation detected.",
                recommendation="None — stale management is configured.",
            )
        return Finding(
            pillar="Operational Excellence",
            principle="Stale issue/PR automation",
            status=Status.FAIL,
            severity=Severity.LOW,
            evidence="No stale issue/PR automation found.",
            recommendation=(
                "Add the `actions/stale` action to auto-label and close stale issues/PRs. "
                "Alternatively, use GitHub's built-in auto-close features or a Probot app."
            ),
            acceptable_risk_note=(
                "For smaller projects or internal repos with few open issues, stale automation "
                "may not be necessary. This is more critical for open-source or high-volume repos."
            ),
        )

    def check_release_automation(self) -> Finding:
        """Check for automated release/versioning workflows."""
        all_wf = self._all_workflow_text()
        has_release = bool(re.search(
            r'semantic.release|release-please|changesets|conventional.commit|auto.*changelog|softprops/action-gh-release',
            all_wf, re.IGNORECASE
        ))
        has_tag_trigger = bool(re.search(r'on:.*push.*tags:', all_wf, re.DOTALL))
        if has_release:
            return Finding(
                pillar="Operational Excellence",
                principle="Automated releases with semantic versioning",
                status=Status.PASS,
                severity=Severity.MEDIUM,
                evidence="Automated release tooling detected in workflows.",
                recommendation="None — release automation is configured.",
            )
        if has_tag_trigger:
            return Finding(
                pillar="Operational Excellence",
                principle="Automated releases with semantic versioning",
                status=Status.PARTIAL,
                severity=Severity.MEDIUM,
                evidence="Tag-triggered workflow found but no semantic release tooling detected.",
                recommendation="Consider adding release-please, semantic-release, or changesets for automated changelog and version management.",
            )
        return Finding(
            pillar="Operational Excellence",
            principle="Automated releases with semantic versioning",
            status=Status.FAIL,
            severity=Severity.MEDIUM,
            evidence="No release automation detected.",
            recommendation=(
                "Implement automated releases. Options: "
                "(1) google-github-actions/release-please-action for conventional commits, "
                "(2) semantic-release for npm projects, "
                "(3) softprops/action-gh-release for tag-based releases."
            ),
            acceptable_risk_note=(
                "Some teams prefer manual release processes for governance reasons. "
                "This is acceptable if documented and the process includes changelog generation."
            ),
        )

    # ==================== PERFORMANCE EFFICIENCY ====================

    def check_dependency_caching(self) -> Finding:
        """Check for dependency caching in workflows."""
        workflows = self._workflow_files()
        if not workflows:
            return Finding(
                pillar="Performance Efficiency",
                principle="Dependency caching in CI/CD",
                status=Status.NOT_APPLICABLE,
                severity=Severity.MEDIUM,
                evidence="No workflows found.",
                recommendation="N/A.",
            )
        all_wf = self._all_workflow_text()
        has_cache = bool(re.search(r'actions/cache|cache:\s*true|with:\s*\n\s*cache', all_wf, re.IGNORECASE))
        # Language-specific caching
        lang_cache = bool(re.search(
            r'actions/setup-node.*cache|actions/setup-python.*cache|actions/setup-java.*cache|actions/setup-go.*cache',
            all_wf, re.IGNORECASE
        ))
        if has_cache or lang_cache:
            return Finding(
                pillar="Performance Efficiency",
                principle="Dependency caching in CI/CD",
                status=Status.PASS,
                severity=Severity.MEDIUM,
                evidence="Dependency caching detected in workflows.",
                recommendation="None — caching is configured.",
            )
        # Check if there's anything to cache
        has_install_step = bool(re.search(r'npm install|pip install|go build|mvn|gradle|yarn|pnpm', all_wf, re.IGNORECASE))
        if has_install_step:
            return Finding(
                pillar="Performance Efficiency",
                principle="Dependency caching in CI/CD",
                status=Status.FAIL,
                severity=Severity.MEDIUM,
                evidence="Package install steps found but no caching configured.",
                recommendation=(
                    "Add dependency caching. Use the built-in `cache` input on setup-* actions, "
                    "or add `actions/cache` with the appropriate key and path. "
                    "This can reduce CI time by 30-70%."
                ),
            )
        return Finding(
            pillar="Performance Efficiency",
            principle="Dependency caching in CI/CD",
            status=Status.NOT_APPLICABLE,
            severity=Severity.MEDIUM,
            evidence="No package install steps detected.",
            recommendation="N/A — no caching opportunity detected.",
        )

    def check_concurrency_groups(self) -> Finding:
        """Check for concurrency groups to cancel redundant runs."""
        workflows = self._workflow_files()
        if not workflows:
            return Finding(
                pillar="Performance Efficiency",
                principle="Concurrency groups for redundant run cancellation",
                status=Status.NOT_APPLICABLE,
                severity=Severity.LOW,
                evidence="No workflows found.",
                recommendation="N/A.",
            )
        all_wf = self._all_workflow_text()
        has_concurrency = bool(re.search(r'concurrency:', all_wf))
        if has_concurrency:
            has_cancel = bool(re.search(r'cancel-in-progress:\s*true', all_wf))
            if has_cancel:
                return Finding(
                    pillar="Performance Efficiency",
                    principle="Concurrency groups for redundant run cancellation",
                    status=Status.PASS,
                    severity=Severity.LOW,
                    evidence="Concurrency groups with cancel-in-progress found.",
                    recommendation="None — concurrency management is configured.",
                )
            return Finding(
                pillar="Performance Efficiency",
                principle="Concurrency groups for redundant run cancellation",
                status=Status.PARTIAL,
                severity=Severity.LOW,
                evidence="Concurrency groups found but cancel-in-progress is not enabled.",
                recommendation="Add `cancel-in-progress: true` to concurrency groups for CI workflows.",
            )
        return Finding(
            pillar="Performance Efficiency",
            principle="Concurrency groups for redundant run cancellation",
            status=Status.FAIL,
            severity=Severity.LOW,
            evidence="No concurrency groups defined.",
            recommendation=(
                "Add concurrency groups to workflows to prevent redundant runs:\n"
                "```yaml\nconcurrency:\n  group: ${{ github.workflow }}-${{ github.ref }}\n"
                "  cancel-in-progress: true\n```"
            ),
        )

    def check_narrow_triggers(self) -> Finding:
        """Check that workflow triggers are scoped narrowly."""
        workflows = self._workflow_files()
        if not workflows:
            return Finding(
                pillar="Performance Efficiency",
                principle="Narrow workflow triggers (paths/branches filters)",
                status=Status.NOT_APPLICABLE,
                severity=Severity.LOW,
                evidence="No workflows found.",
                recommendation="N/A.",
            )
        broad_triggers = []
        filtered_count = 0
        for wf in workflows:
            text = read_text(wf)
            name = str(wf.relative_to(self.repo))
            has_path_filter = bool(re.search(r'paths:', text))
            has_branch_filter = bool(re.search(r'branches:', text))
            if has_path_filter or has_branch_filter:
                filtered_count += 1
            elif re.search(r'on:\s*\[?\s*push', text):
                broad_triggers.append(name)
        if not broad_triggers:
            return Finding(
                pillar="Performance Efficiency",
                principle="Narrow workflow triggers (paths/branches filters)",
                status=Status.PASS,
                severity=Severity.LOW,
                evidence=f"All workflows use branch or path filters.",
                recommendation="None — triggers are well-scoped.",
            )
        if filtered_count > 0:
            return Finding(
                pillar="Performance Efficiency",
                principle="Narrow workflow triggers (paths/branches filters)",
                status=Status.PARTIAL,
                severity=Severity.LOW,
                evidence=f"Some workflows lack path/branch filters: {', '.join(broad_triggers[:3])}",
                recommendation="Add `paths:` or `branches:` filters to avoid unnecessary workflow runs.",
                acceptable_risk_note=(
                    "Small monorepos or single-purpose repos may intentionally run CI on all pushes. "
                    "This is acceptable when the CI is fast and the repo scope is well-defined."
                ),
            )
        return Finding(
            pillar="Performance Efficiency",
            principle="Narrow workflow triggers (paths/branches filters)",
            status=Status.FAIL,
            severity=Severity.LOW,
            evidence="No workflows use path or branch filters.",
            recommendation="Add `paths:` and `branches:` filters to reduce unnecessary workflow runs.",
        )

    def check_matrix_strategies(self) -> Finding:
        """Check for matrix testing strategies."""
        all_wf = self._all_workflow_text()
        if not all_wf:
            return Finding(
                pillar="Performance Efficiency",
                principle="Matrix strategies for parallel testing",
                status=Status.NOT_APPLICABLE,
                severity=Severity.LOW,
                evidence="No workflows found.",
                recommendation="N/A.",
            )
        has_matrix = bool(re.search(r'strategy:\s*\n\s*matrix:', all_wf))
        if has_matrix:
            return Finding(
                pillar="Performance Efficiency",
                principle="Matrix strategies for parallel testing",
                status=Status.PASS,
                severity=Severity.LOW,
                evidence="Matrix strategy detected in workflows.",
                recommendation="None — parallel testing is configured.",
            )
        has_test = bool(re.search(r'test|spec|lint|check', all_wf, re.IGNORECASE))
        if has_test:
            return Finding(
                pillar="Performance Efficiency",
                principle="Matrix strategies for parallel testing",
                status=Status.PARTIAL,
                severity=Severity.LOW,
                evidence="Test/check steps found but no matrix strategy for parallel execution.",
                recommendation=(
                    "Consider using matrix strategies for cross-version or cross-platform testing:\n"
                    "```yaml\nstrategy:\n  matrix:\n    node-version: [18, 20, 22]\n```"
                ),
                acceptable_risk_note=(
                    "Single-platform, single-version projects may not need matrix testing. "
                    "This is most valuable when supporting multiple runtime versions or OS platforms."
                ),
            )
        return Finding(
            pillar="Performance Efficiency",
            principle="Matrix strategies for parallel testing",
            status=Status.NOT_APPLICABLE,
            severity=Severity.LOW,
            evidence="No test steps detected in workflows.",
            recommendation="N/A.",
        )

    # ==================== COLLABORATION ====================

    def check_codeowners(self) -> Finding:
        """Check for CODEOWNERS file."""
        path = find_file(self.repo, ".github/CODEOWNERS", "CODEOWNERS", "docs/CODEOWNERS")
        if path:
            text = read_text(path)
            rules = [l for l in text.splitlines() if l.strip() and not l.strip().startswith("#")]
            if rules:
                return Finding(
                    pillar="Collaboration",
                    principle="CODEOWNERS for automatic review assignment",
                    status=Status.PASS,
                    severity=Severity.MEDIUM,
                    evidence=f"CODEOWNERS found with {len(rules)} rule(s).",
                    recommendation="None — CODEOWNERS is configured.",
                )
            return Finding(
                pillar="Collaboration",
                principle="CODEOWNERS for automatic review assignment",
                status=Status.PARTIAL,
                severity=Severity.MEDIUM,
                evidence="CODEOWNERS exists but contains no rules.",
                recommendation="Add ownership rules mapping paths to teams/individuals.",
            )
        return Finding(
            pillar="Collaboration",
            principle="CODEOWNERS for automatic review assignment",
            status=Status.FAIL,
            severity=Severity.MEDIUM,
            evidence="No CODEOWNERS file found.",
            recommendation=(
                "Create .github/CODEOWNERS to automate review assignment. "
                "Example: `* @org/engineering-team`"
            ),
        )

    def check_copilot_instructions(self) -> Finding:
        """Check for Copilot custom instructions."""
        candidates = [
            ".github/copilot-instructions.md",
            ".github/instructions",  # directory
        ]
        has_instructions = find_file(self.repo, ".github/copilot-instructions.md")
        instructions_dir = self.repo / ".github" / "instructions"
        has_dir = instructions_dir.is_dir() and any(instructions_dir.iterdir())

        if has_instructions or has_dir:
            return Finding(
                pillar="Collaboration",
                principle="Copilot custom instructions",
                status=Status.PASS,
                severity=Severity.LOW,
                evidence=f"Copilot instructions found: {'copilot-instructions.md' if has_instructions else 'instructions directory'}",
                recommendation="None — Copilot is customized for the project.",
            )
        return Finding(
            pillar="Collaboration",
            principle="Copilot custom instructions",
            status=Status.FAIL,
            severity=Severity.LOW,
            evidence="No Copilot custom instructions found.",
            recommendation=(
                "Create .github/copilot-instructions.md with project-specific coding standards, "
                "architecture guidelines, and conventions for Copilot to follow."
            ),
            acceptable_risk_note=(
                "Not all organizations use GitHub Copilot. If Copilot is not part of your toolchain, "
                "this is not applicable. However, customization files also benefit future adoption."
            ),
        )

    def check_issue_templates(self) -> Finding:
        """Check for issue templates."""
        template_dir = self.repo / ".github" / "ISSUE_TEMPLATE"
        has_template_dir = template_dir.is_dir() and any(template_dir.iterdir())
        has_single = find_file(self.repo, ".github/ISSUE_TEMPLATE.md", "ISSUE_TEMPLATE.md")
        if has_template_dir or has_single:
            return Finding(
                pillar="Collaboration",
                principle="Issue templates for standardized reporting",
                status=Status.PASS,
                severity=Severity.LOW,
                evidence="Issue template(s) found.",
                recommendation="None — issue templates are configured.",
            )
        return Finding(
            pillar="Collaboration",
            principle="Issue templates for standardized reporting",
            status=Status.FAIL,
            severity=Severity.LOW,
            evidence="No issue templates found.",
            recommendation=(
                "Create .github/ISSUE_TEMPLATE/ with templates for bugs, features, etc. "
                "Use YAML-based issue forms for structured input."
            ),
        )

    def check_license(self) -> Finding:
        """Check for LICENSE file."""
        path = find_file(self.repo, "LICENSE", "LICENSE.md", "LICENSE.txt", "LICENCE", "COPYING")
        if path:
            return Finding(
                pillar="Collaboration",
                principle="License file present",
                status=Status.PASS,
                severity=Severity.MEDIUM,
                evidence=f"License found: {path.relative_to(self.repo)}",
                recommendation="None — license is in place.",
            )
        return Finding(
            pillar="Collaboration",
            principle="License file present",
            status=Status.FAIL,
            severity=Severity.MEDIUM,
            evidence="No LICENSE file found.",
            recommendation="Add a LICENSE file. For open-source, choose from https://choosealicense.com/. For internal projects, specify 'All Rights Reserved' or your org's standard license.",
        )


# ---------------------------------------------------------------------------
# Interactive organizational assessment questions
# ---------------------------------------------------------------------------

ORG_QUESTIONS: list[dict] = [
    # ── Security ──
    {
        "pillar": "Security",
        "principle": "Secret scanning enabled at org level",
        "severity": Severity.CRITICAL,
        "question": "Is GitHub Secret Scanning enabled for your organization?",
        "type": "choice",
        "options": [
            "Yes — enabled for all repositories",
            "Partially — enabled for some repositories",
            "No — not enabled",
            "Don't know",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.FAIL, Status.FAIL],
        "recommendation": (
            "Enable secret scanning at the organization level: "
            "Settings → Code security → Secret scanning → Enable for all repos. "
            "Also enable push protection to block secrets before they're committed."
        ),
        "acceptable_risk": "",
    },
    {
        "pillar": "Security",
        "principle": "Push protection for secrets",
        "severity": Severity.CRITICAL,
        "question": "Is push protection enabled (blocks commits containing secrets before they reach GitHub)?",
        "type": "choice",
        "options": [
            "Yes — enabled org-wide",
            "Yes — enabled for some repos",
            "No — not enabled",
            "Don't know",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.FAIL, Status.FAIL],
        "recommendation": (
            "Enable push protection at the org level to prevent secrets from being pushed. "
            "Settings → Code security → Push protection. "
            "Users can bypass with a reason, which is auditable."
        ),
        "acceptable_risk": (
            "Some CI/CD secrets may trigger false positives. "
            "Push protection allows bypass with documented reasons — this is acceptable "
            "as long as bypass events are reviewed regularly."
        ),
    },
    {
        "pillar": "Security",
        "principle": "Private vulnerability reporting enabled",
        "severity": Severity.HIGH,
        "question": "Is private vulnerability reporting enabled for your repositories?",
        "type": "choice",
        "options": [
            "Yes — enabled org-wide",
            "Yes — enabled for some repos",
            "No",
            "Don't know",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.FAIL, Status.FAIL],
        "recommendation": (
            "Enable private vulnerability reporting so security researchers can report issues "
            "confidentially. Settings → Code security → Private vulnerability reporting."
        ),
        "acceptable_risk": (
            "Internal-only repositories without external contribution may have lower priority "
            "for this feature, but it's still recommended as a defense-in-depth measure."
        ),
    },
    # ── Reliability ──
    {
        "pillar": "Reliability",
        "principle": "Branch protection on default branch",
        "severity": Severity.CRITICAL,
        "question": "Is branch protection enabled on the default branch (main/master)?",
        "type": "choice",
        "options": [
            "Yes — with required reviews AND required status checks",
            "Partially — required reviews OR status checks (not both)",
            "Minimal — branch protection enabled but no required checks",
            "No — no branch protection",
            "Using rulesets instead of classic branch protection",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.PARTIAL, Status.FAIL, Status.PASS],
        "recommendation": (
            "Enable branch protection (or rulesets) on the default branch requiring: "
            "(1) Pull request reviews before merge, (2) Required status checks (CI must pass), "
            "(3) Up-to-date branches before merge. "
            "See: https://docs.github.com/en/repositories/configuring-branches-and-merges-in-your-repository"
        ),
        "acceptable_risk": "",
        "alternative": (
            "GitHub Rulesets are the modern replacement for classic branch protection rules. "
            "They offer org-level management, bypass permissions, and more granular controls. "
            "If using rulesets, classic branch protection is not needed."
        ),
    },
    {
        "pillar": "Reliability",
        "principle": "Signed commits required",
        "severity": Severity.MEDIUM,
        "question": "Are signed commits required on protected branches?",
        "type": "choice",
        "options": [
            "Yes — required via branch protection/rulesets",
            "Encouraged but not enforced",
            "No",
            "Don't know",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.FAIL, Status.FAIL],
        "recommendation": (
            "Require signed commits on protected branches to verify commit author identity. "
            "Enable in branch protection settings or rulesets."
        ),
        "acceptable_risk": (
            "Commit signing adds friction and requires GPG/SSH key setup for all contributors. "
            "Many organizations accept unsigned commits as a known risk if they have strong "
            "authentication (SSO/MFA) and branch protection requiring PR reviews."
        ),
    },
    {
        "pillar": "Reliability",
        "principle": "Required number of reviewers",
        "severity": Severity.HIGH,
        "question": "How many reviewers are required before merging a pull request?",
        "type": "choice",
        "options": [
            "2 or more reviewers required",
            "1 reviewer required",
            "Reviews not required (informational only)",
            "Don't know / varies by repo",
        ],
        "status_map": [Status.PASS, Status.PASS, Status.FAIL, Status.PARTIAL],
        "recommendation": (
            "Require at least 1 reviewer (2 for high-risk repos like production infrastructure). "
            "Configure in branch protection: Require pull request reviews → Required approving reviews."
        ),
        "acceptable_risk": (
            "1 reviewer is acceptable for most repos. 2+ reviewers is recommended for: "
            "production infrastructure, security-sensitive code, and compliance-regulated projects. "
            "Small teams (1-2 developers) may need to approve their own PRs — document this exception."
        ),
    },
    # ── Operational Excellence ──
    {
        "pillar": "Operational Excellence",
        "principle": "Work tracking with GitHub Projects/Issues",
        "severity": Severity.MEDIUM,
        "question": "How does your team track work and plan sprints?",
        "type": "choice",
        "options": [
            "GitHub Projects (new) — boards, tables, and roadmaps",
            "GitHub Issues only (no project board)",
            "External tool (Jira, Azure DevOps, Linear, etc.)",
            "No formal work tracking",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.ALTERNATIVE, Status.FAIL],
        "recommendation": (
            "Use GitHub Projects for work tracking to keep planning close to code. "
            "Projects v2 supports boards, tables, roadmaps, and custom fields."
        ),
        "acceptable_risk": "",
        "alternative": (
            "External tools like Jira, Azure DevOps, or Linear are acceptable alternatives, "
            "especially if your org has an existing investment. Ensure 2-way sync with GitHub "
            "Issues using integrations (e.g., Jira for GitHub, Azure DevOps GitHub sync) "
            "so PRs link to work items. The key principle is traceability, not the specific tool."
        ),
    },
    {
        "pillar": "Operational Excellence",
        "principle": "InnerSource practices for cross-team collaboration",
        "severity": Severity.LOW,
        "question": "Does your organization practice InnerSource (internal open-source practices)?",
        "type": "choice",
        "options": [
            "Yes — we have documented InnerSource guidelines and active cross-team contributions",
            "Partially — some cross-team contributions happen informally",
            "No — teams work in silos",
            "Not applicable (single team / small org)",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.FAIL, Status.NOT_APPLICABLE],
        "recommendation": (
            "Adopt InnerSource practices: make repos discoverable with good READMEs, "
            "accept contributions from other teams via PRs, use CONTRIBUTING.md to set expectations, "
            "and recognize cross-team contributors."
        ),
        "acceptable_risk": (
            "Single-team or very small organizations may not benefit from formal InnerSource. "
            "The principles still apply at smaller scale — clear docs, welcoming contribution, and discoverability."
        ),
    },
    # ── Performance Efficiency ──
    {
        "pillar": "Performance Efficiency",
        "principle": "CI/CD pipeline run time monitoring",
        "severity": Severity.MEDIUM,
        "question": "Do you actively monitor and optimize CI/CD workflow run times?",
        "type": "choice",
        "options": [
            "Yes — we track metrics and have optimization targets",
            "Sometimes — we investigate when builds are slow",
            "No — we don't track CI/CD performance",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.FAIL],
        "recommendation": (
            "Monitor workflow run duration using GitHub Actions usage metrics, "
            "or third-party tools like BuildPulse, Datadog CI, or Trunk. "
            "Set targets (e.g., CI < 10 minutes) and investigate regressions."
        ),
        "acceptable_risk": "",
    },
    {
        "pillar": "Performance Efficiency",
        "principle": "Self-hosted or larger runners for compute-intensive jobs",
        "severity": Severity.LOW,
        "question": "Do you use self-hosted or larger GitHub-hosted runners for compute-intensive jobs?",
        "type": "choice",
        "options": [
            "Yes — we use larger runners or self-hosted runners",
            "No — standard GitHub-hosted runners only",
            "Not applicable — our builds are lightweight",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.NOT_APPLICABLE],
        "recommendation": (
            "For compute-intensive builds (large compilations, ML training, heavy Docker builds), "
            "consider GitHub-hosted larger runners or self-hosted runners to reduce build times."
        ),
        "acceptable_risk": (
            "Standard runners are adequate for most projects. Larger runners add cost "
            "and complexity. Only invest when build times are a measurable bottleneck."
        ),
    },
    # ── Collaboration ──
    {
        "pillar": "Collaboration",
        "principle": "GitHub Discussions or async communication channel",
        "severity": Severity.LOW,
        "question": "Is GitHub Discussions (or equivalent async forum) enabled for your repositories?",
        "type": "choice",
        "options": [
            "Yes — GitHub Discussions is enabled",
            "We use an alternative (Slack channels, Teams, Discourse, etc.)",
            "No async discussion forum",
        ],
        "status_map": [Status.PASS, Status.ALTERNATIVE, Status.FAIL],
        "recommendation": (
            "Enable GitHub Discussions for Q&A, design proposals, and community engagement. "
            "Settings → General → Features → Discussions."
        ),
        "acceptable_risk": "",
        "alternative": (
            "Slack, Microsoft Teams, or Discourse are acceptable alternatives for async communication. "
            "The key principle is having a dedicated, searchable space for discussions "
            "that's separate from issue tracking."
        ),
    },
    {
        "pillar": "Collaboration",
        "principle": "Branch naming conventions enforced",
        "severity": Severity.LOW,
        "question": "Does your org enforce branch naming conventions (e.g., feature/, bugfix/, release/)?",
        "type": "choice",
        "options": [
            "Yes — enforced via rulesets or automation",
            "Documented but not enforced",
            "No conventions",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.FAIL],
        "recommendation": (
            "Define and enforce branch naming conventions using GitHub Rulesets at the org level. "
            "Example pattern: `feature/*`, `bugfix/*`, `release/*`. "
            "This improves clarity and enables automated workflows based on branch patterns."
        ),
        "acceptable_risk": (
            "Trunk-based development with very short-lived branches may not benefit from "
            "strict naming conventions. The key is having a clear, documented branching strategy."
        ),
    },
    {
        "pillar": "Collaboration",
        "principle": "Onboarding automation for new contributors",
        "severity": Severity.LOW,
        "question": "Do you have onboarding automation for new repository contributors?",
        "type": "choice",
        "options": [
            "Yes — automated welcome messages, guided setup, and docs",
            "Partially — we have docs but no automation",
            "No — contributors figure it out themselves",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.FAIL],
        "recommendation": (
            "Consider: (1) A GitHub Action to welcome first-time contributors, "
            "(2) A devcontainer.json for one-click dev environment setup, "
            "(3) A 'Getting Started' section in CONTRIBUTING.md, "
            "(4) GitHub Codespaces for instant development environments."
        ),
        "acceptable_risk": (
            "For small internal teams with stable membership, formal onboarding automation "
            "may be overkill. Focus on clear documentation at minimum."
        ),
    },
    # ── Org-level Security (additional) ──
    {
        "pillar": "Security",
        "principle": "Organization SSO and MFA enforcement",
        "severity": Severity.CRITICAL,
        "question": "Does your GitHub organization enforce SSO (SAML) and/or require MFA for all members?",
        "type": "choice",
        "options": [
            "Yes — SAML SSO enforced with MFA",
            "MFA required but no SSO",
            "SSO configured but MFA not enforced",
            "Neither SSO nor MFA enforced",
            "Not applicable (personal account / no org)",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.PARTIAL, Status.FAIL, Status.NOT_APPLICABLE],
        "recommendation": (
            "Enforce SAML SSO via your identity provider (Okta, Azure AD, etc.) and require MFA. "
            "Settings → Authentication security → Require two-factor authentication. "
            "For Enterprise: Settings → Authentication → SAML single sign-on."
        ),
        "acceptable_risk": (
            "Small teams or open-source projects may not need SAML SSO. "
            "At minimum, require MFA for all organization members — this is non-negotiable for security."
        ),
    },
    {
        "pillar": "Security",
        "principle": "GitHub Advanced Security (GHAS) licensing",
        "severity": Severity.HIGH,
        "question": "Does your organization have GitHub Advanced Security (GHAS) licenses?",
        "type": "choice",
        "options": [
            "Yes — GHAS enabled for all repos",
            "Yes — GHAS enabled for some repos (prioritized)",
            "No — using free tier security features only",
            "Don't know",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.PARTIAL, Status.FAIL],
        "recommendation": (
            "GHAS provides CodeQL, secret scanning with custom patterns, dependency review, "
            "and security overview dashboards. If budget is limited, prioritize GHAS for "
            "production and customer-facing repositories."
        ),
        "acceptable_risk": (
            "GHAS has a per-committer cost. Organizations can achieve baseline security with "
            "free features (Dependabot, basic secret scanning for public repos) plus third-party "
            "SAST tools. Prioritize GHAS for your most critical repositories."
        ),
    },
    # ── Reliability (additional) ──
    {
        "pillar": "Reliability",
        "principle": "Disaster recovery / backup strategy for GitHub data",
        "severity": Severity.HIGH,
        "question": "Do you have a backup/disaster recovery strategy for your GitHub repositories and metadata?",
        "type": "choice",
        "options": [
            "Yes — automated backups of repos, issues, wikis, and settings",
            "Partially — Git repos are backed up but not metadata (issues, PRs, settings)",
            "No — relying on GitHub's infrastructure",
            "Don't know",
        ],
        "status_map": [Status.PASS, Status.PARTIAL, Status.PARTIAL, Status.FAIL],
        "recommendation": (
            "While GitHub has robust infrastructure, consider: "
            "(1) Regular git mirror backups to another provider, "
            "(2) GitHub API exports of issues, PRs, and wiki content, "
            "(3) Tools like BackHub, GitHub Backup Utils (for GHES), or custom scripts. "
            "This is especially important for compliance-regulated environments."
        ),
        "acceptable_risk": (
            "GitHub's enterprise SLA provides strong uptime guarantees. Many organizations "
            "accept the risk of relying solely on GitHub's infrastructure. This is generally "
            "acceptable for non-regulated environments, but consider at least mirroring "
            "critical repositories."
        ),
    },
]


# ---------------------------------------------------------------------------
# Assessment engine
# ---------------------------------------------------------------------------

class WellArchitectedAssessment:
    """Orchestrates the full assessment."""

    def __init__(self, repo_path: Path, skip_interactive: bool = False):
        self.repo_path = repo_path
        self.skip_interactive = skip_interactive
        self.scanner = RepoScanner(repo_path)
        self.pillars: dict[str, PillarScore] = {
            "Security": PillarScore("Security"),
            "Reliability": PillarScore("Reliability"),
            "Operational Excellence": PillarScore("Operational Excellence"),
            "Performance Efficiency": PillarScore("Performance Efficiency"),
            "Collaboration": PillarScore("Collaboration"),
        }

    def _add(self, finding: Finding) -> None:
        self.pillars[finding.pillar].findings.append(finding)

    def run_automated_checks(self) -> None:
        """Run all file-system-based automated checks."""
        section("Running automated repository scans...")

        # Security
        checks = [
            ("Dependabot config", self.scanner.check_dependabot_config),
            ("CodeQL / code scanning", self.scanner.check_codeql_scanning),
            ("SECURITY.md", self.scanner.check_security_md),
            ("GITHUB_TOKEN permissions", self.scanner.check_token_permissions),
            ("Action SHA pinning", self.scanner.check_action_pinning),
            ("OIDC for cloud auth", self.scanner.check_oidc_usage),
        ]
        for label, check_fn in checks:
            finding = check_fn()
            self._add(finding)
            icon = status_icon(finding.status)
            print(f"  {icon} {label}: {finding.status.value}")

        # Reliability
        checks = [
            ("CI workflows", self.scanner.check_ci_workflows),
            ("Rollback strategy", self.scanner.check_rollback_strategy),
            ("Deployment environments", self.scanner.check_environments),
        ]
        for label, check_fn in checks:
            finding = check_fn()
            self._add(finding)
            icon = status_icon(finding.status)
            print(f"  {icon} {label}: {finding.status.value}")

        # Operational Excellence
        checks = [
            ("CONTRIBUTING.md", self.scanner.check_contributing_md),
            ("PR template", self.scanner.check_pr_template),
            ("Reusable workflows", self.scanner.check_reusable_workflows),
            ("README quality", self.scanner.check_readme_quality),
            ("Stale automation", self.scanner.check_stale_automation),
            ("Release automation", self.scanner.check_release_automation),
        ]
        for label, check_fn in checks:
            finding = check_fn()
            self._add(finding)
            icon = status_icon(finding.status)
            print(f"  {icon} {label}: {finding.status.value}")

        # Performance Efficiency
        checks = [
            ("Dependency caching", self.scanner.check_dependency_caching),
            ("Concurrency groups", self.scanner.check_concurrency_groups),
            ("Narrow triggers", self.scanner.check_narrow_triggers),
            ("Matrix strategies", self.scanner.check_matrix_strategies),
        ]
        for label, check_fn in checks:
            finding = check_fn()
            self._add(finding)
            icon = status_icon(finding.status)
            print(f"  {icon} {label}: {finding.status.value}")

        # Collaboration
        checks = [
            ("CODEOWNERS", self.scanner.check_codeowners),
            ("Copilot instructions", self.scanner.check_copilot_instructions),
            ("Issue templates", self.scanner.check_issue_templates),
            ("License", self.scanner.check_license),
        ]
        for label, check_fn in checks:
            finding = check_fn()
            self._add(finding)
            icon = status_icon(finding.status)
            print(f"  {icon} {label}: {finding.status.value}")

    def run_interactive_questions(self) -> None:
        """Ask organizational-level questions that can't be detected from files."""
        if self.skip_interactive:
            info("Skipping interactive questions (--no-interactive mode).")
            return

        section("Organizational Policy Assessment")
        info("The following questions cover settings that can't be detected from files alone.")
        info("These are typically configured at the organization or repository settings level.\n")

        current_pillar = ""
        for q in ORG_QUESTIONS:
            if q["pillar"] != current_pillar:
                current_pillar = q["pillar"]
                print(f"\n  {_c(f'─── {current_pillar} ───', 'bold')}")

            idx = ask_choice(q["question"], q["options"])
            status = q["status_map"][idx]
            evidence = f"User response: {q['options'][idx]}"

            finding = Finding(
                pillar=q["pillar"],
                principle=q["principle"],
                status=status,
                severity=q["severity"],
                evidence=evidence,
                recommendation=q["recommendation"] if status != Status.PASS else "None — meets the standard.",
                acceptable_risk_note=q.get("acceptable_risk", ""),
                alternative_note=q.get("alternative", ""),
            )
            self._add(finding)

    def run(self) -> str:
        """Execute the full assessment and return the markdown report."""
        banner("GitHub Well-Architected Framework Assessment")
        info(f"Repository: {self.repo_path}")
        info(f"Date: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
        info(f"Framework: {FRAMEWORK_URL}")

        self.run_automated_checks()
        self.run_interactive_questions()

        report = self._generate_report()

        # Print summary to terminal
        section("Assessment Complete — Summary")
        total = sum(p.score for p in self.pillars.values())
        for name, pillar in self.pillars.items():
            color = "green" if pillar.score >= 8 else ("yellow" if pillar.score >= 5 else "red")
            print(f"  {_c(f'{pillar.score:4.1f}/10', color)}  {name} — {pillar.compliance}")
        print()
        color = "green" if total >= 40 else ("yellow" if total >= 25 else "red")
        print(f"  {_c(f'Overall: {total:.1f}/50', color)}")
        print()

        return report

    def _generate_report(self) -> str:
        """Generate the full markdown compliance report."""
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        total = sum(p.score for p in self.pillars.values())

        # Executive summary
        fail_count = sum(
            1 for p in self.pillars.values()
            for f in p.findings if f.status == Status.FAIL
        )
        critical_fails = sum(
            1 for p in self.pillars.values()
            for f in p.findings if f.status == Status.FAIL and f.severity == Severity.CRITICAL
        )
        alt_count = sum(
            1 for p in self.pillars.values()
            for f in p.findings
            if f.status in (Status.ACCEPTABLE_RISK, Status.ALTERNATIVE) or f.alternative_note
        )

        if total >= 40:
            exec_summary = f"This repository demonstrates **strong alignment** with the GitHub Well-Architected Framework, scoring {total:.1f}/50. "
        elif total >= 25:
            exec_summary = f"This repository has **partial alignment** with the GitHub Well-Architected Framework, scoring {total:.1f}/50. "
        else:
            exec_summary = f"This repository has **significant gaps** relative to the GitHub Well-Architected Framework, scoring {total:.1f}/50. "

        if fail_count > 0:
            exec_summary += f"There are **{fail_count} gaps** requiring attention"
            if critical_fails > 0:
                exec_summary += f" ({critical_fails} critical)"
            exec_summary += ". "
        if alt_count > 0:
            exec_summary += f"**{alt_count} area(s)** use acceptable alternatives or carry documented acceptable risk."

        lines = [
            "# GitHub Well-Architected Framework — Compliance Report",
            "",
            f"**Repository**: `{self.repo_path.name}`  ",
            f"**Date**: {now}  ",
            f"**Framework**: [{FRAMEWORK_URL}]({FRAMEWORK_URL})  ",
            f"**Assessment Method**: Automated scanning + interactive organizational review  ",
            "",
            "---",
            "",
            "## Executive Summary",
            "",
            exec_summary,
            "",
            "| Pillar | Score | Compliance |",
            "|--------|------:|------------|",
        ]
        for name, pillar in self.pillars.items():
            lines.append(f"| {name} | {pillar.score:.1f}/10 | {pillar.compliance} |")
        lines.append(f"| **Overall** | **{total:.1f}/50** | |")
        lines.append("")
        lines.append("---")
        lines.append("")

        # Detailed pillar sections
        for name, pillar in self.pillars.items():
            lines.append(f"## {name} — {pillar.compliance} ({pillar.score:.1f}/10)")
            lines.append("")
            lines.append("| # | Principle | Status | Severity | Evidence | Recommendation |")
            lines.append("|---|-----------|--------|----------|----------|----------------|")
            for i, f in enumerate(pillar.findings, 1):
                status_emoji = {
                    Status.PASS: "✅",
                    Status.PARTIAL: "🟡",
                    Status.FAIL: "❌",
                    Status.ACCEPTABLE_RISK: "⚡",
                    Status.ALTERNATIVE: "↔️",
                    Status.NOT_APPLICABLE: "➖",
                }.get(f.status, "❓")
                rec = f.recommendation.replace("\n", " ").replace("|", "\\|")
                ev = f.evidence.replace("\n", " ").replace("|", "\\|")
                lines.append(
                    f"| {i} | {f.principle} | {status_emoji} {f.status.value} | {f.severity.value} | {ev} | {rec} |"
                )
            lines.append("")

            # Acceptable risks and alternatives for this pillar
            risk_findings = [f for f in pillar.findings if f.acceptable_risk_note]
            alt_findings = [f for f in pillar.findings if f.alternative_note]
            if risk_findings or alt_findings:
                lines.append(f"### {name} — Notes on Acceptable Risks & Alternatives")
                lines.append("")
                for f in risk_findings:
                    lines.append(f"- **{f.principle}** — ⚡ *Acceptable Risk*: {f.acceptable_risk_note}")
                for f in alt_findings:
                    lines.append(f"- **{f.principle}** — ↔️ *Alternative Approach*: {f.alternative_note}")
                lines.append("")

            lines.append("---")
            lines.append("")

        # Priority actions
        lines.append("## Priority Actions")
        lines.append("")
        lines.append("The following items are ordered by severity and impact:")
        lines.append("")
        # Collect all non-passing, non-N/A findings and sort by severity
        action_items = []
        for p in self.pillars.values():
            for f in p.findings:
                if f.status in (Status.FAIL, Status.PARTIAL):
                    action_items.append(f)
        severity_order = {Severity.CRITICAL: 0, Severity.HIGH: 1, Severity.MEDIUM: 2, Severity.LOW: 3, Severity.INFO: 4}
        status_order = {Status.FAIL: 0, Status.PARTIAL: 1}
        action_items.sort(key=lambda f: (severity_order.get(f.severity, 9), status_order.get(f.status, 9)))

        for i, f in enumerate(action_items, 1):
            status_emoji = "❌" if f.status == Status.FAIL else "🟡"
            lines.append(f"{i}. {status_emoji} **[{f.severity.value}]** [{f.pillar}] {f.principle}")
            lines.append(f"   - {f.recommendation}")
            if f.acceptable_risk_note:
                lines.append(f"   - ⚡ *Acceptable risk*: {f.acceptable_risk_note}")
            if f.alternative_note:
                lines.append(f"   - ↔️ *Alternative*: {f.alternative_note}")
            lines.append("")

        if not action_items:
            lines.append("*No priority actions — all checks pass! 🎉*")
            lines.append("")

        # Gap summary with categorization
        lines.append("---")
        lines.append("")
        lines.append("## Gap Analysis Summary")
        lines.append("")
        gap_categories = {
            "Critical Gaps (must fix)": [f for f in action_items if f.severity == Severity.CRITICAL and f.status == Status.FAIL],
            "High-Priority Gaps": [f for f in action_items if f.severity == Severity.HIGH and f.status == Status.FAIL],
            "Medium-Priority Improvements": [f for f in action_items if f.severity == Severity.MEDIUM],
            "Low-Priority / Nice-to-Have": [f for f in action_items if f.severity in (Severity.LOW, Severity.INFO)],
            "Partial Compliance (needs strengthening)": [f for f in action_items if f.status == Status.PARTIAL and f.severity in (Severity.CRITICAL, Severity.HIGH)],
        }
        for category, items in gap_categories.items():
            if items:
                lines.append(f"### {category}")
                lines.append("")
                for f in items:
                    lines.append(f"- **{f.principle}** ({f.pillar}): {f.evidence}")
                lines.append("")

        # Acceptable risks summary
        all_risks = [
            f for p in self.pillars.values()
            for f in p.findings
            if f.acceptable_risk_note
        ]
        all_alts = [
            f for p in self.pillars.values()
            for f in p.findings
            if f.alternative_note
        ]
        if all_risks or all_alts:
            lines.append("---")
            lines.append("")
            lines.append("## Acceptable Risks & Alternative Approaches")
            lines.append("")
            lines.append("The following findings represent areas where the standard recommendation may not be the only valid approach:")
            lines.append("")
            if all_risks:
                lines.append("### Acceptable Risks")
                lines.append("")
                lines.append("| Pillar | Principle | Risk Description |")
                lines.append("|--------|-----------|-----------------|")
                for f in all_risks:
                    lines.append(f"| {f.pillar} | {f.principle} | {f.acceptable_risk_note.replace(chr(10), ' ')} |")
                lines.append("")
            if all_alts:
                lines.append("### Alternative Approaches")
                lines.append("")
                lines.append("| Pillar | Principle | Alternative |")
                lines.append("|--------|-----------|------------|")
                for f in all_alts:
                    lines.append(f"| {f.pillar} | {f.principle} | {f.alternative_note.replace(chr(10), ' ')} |")
                lines.append("")

        # Footer
        lines.append("---")
        lines.append("")
        lines.append("*Generated by the GitHub Well-Architected Assessment Tool*  ")
        lines.append(f"*Framework reference: [{FRAMEWORK_URL}]({FRAMEWORK_URL})*")
        lines.append("")

        return "\n".join(lines)


# ---------------------------------------------------------------------------
# HTML report generator
# ---------------------------------------------------------------------------

def generate_html_report(pillars: dict[str, PillarScore], repo_name: str) -> str:
    """Generate a modern, self-contained HTML dashboard."""
    h = html_mod.escape
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    total = sum(p.score for p in pillars.values())
    total_pct = round(total / 50 * 100)

    # Pre-compute stats
    all_findings = [f for p in pillars.values() for f in p.findings]
    pass_count = sum(1 for f in all_findings if f.status == Status.PASS)
    partial_count = sum(1 for f in all_findings if f.status in (Status.PARTIAL, Status.ACCEPTABLE_RISK, Status.ALTERNATIVE))
    fail_count = sum(1 for f in all_findings if f.status == Status.FAIL)
    na_count = sum(1 for f in all_findings if f.status == Status.NOT_APPLICABLE)

    # Pillar data for chart
    pillar_names_json = json.dumps([n for n in pillars])
    pillar_scores_json = json.dumps([p.score for p in pillars.values()])

    # Color helpers
    def score_color(s: float) -> str:
        if s >= 8: return "#10b981"
        if s >= 5: return "#f59e0b"
        return "#ef4444"

    def status_badge(s: Status) -> str:
        colors = {
            Status.PASS: ("#10b981", "#ecfdf5", "✓ Pass"),
            Status.PARTIAL: ("#f59e0b", "#fffbeb", "◐ Partial"),
            Status.FAIL: ("#ef4444", "#fef2f2", "✗ Fail"),
            Status.ACCEPTABLE_RISK: ("#8b5cf6", "#f5f3ff", "⚡ Acceptable Risk"),
            Status.ALTERNATIVE: ("#3b82f6", "#eff6ff", "↔ Alternative"),
            Status.NOT_APPLICABLE: ("#6b7280", "#f9fafb", "— N/A"),
        }
        fg, bg, label = colors.get(s, ("#6b7280", "#f9fafb", "?"))
        return f'<span class="badge" style="background:{bg};color:{fg};border:1px solid {fg}20">{label}</span>'

    def severity_badge(sev: Severity) -> str:
        colors = {
            Severity.CRITICAL: ("#dc2626", "#fef2f2"),
            Severity.HIGH: ("#ea580c", "#fff7ed"),
            Severity.MEDIUM: ("#d97706", "#fffbeb"),
            Severity.LOW: ("#2563eb", "#eff6ff"),
            Severity.INFO: ("#6b7280", "#f9fafb"),
        }
        fg, bg = colors.get(sev, ("#6b7280", "#f9fafb"))
        return f'<span class="badge" style="background:{bg};color:{fg};border:1px solid {fg}20">{sev.value}</span>'

    # Build pillar sections
    pillar_html_sections = []
    for pname, pillar in pillars.items():
        sc = pillar.score
        comp = pillar.compliance
        color = score_color(sc)

        rows = []
        for i, f in enumerate(pillar.findings, 1):
            notes_parts = []
            if f.acceptable_risk_note:
                notes_parts.append(f'<div class="note note-risk"><strong>⚡ Acceptable Risk:</strong> {h(f.acceptable_risk_note)}</div>')
            if f.alternative_note:
                notes_parts.append(f'<div class="note note-alt"><strong>↔ Alternative:</strong> {h(f.alternative_note)}</div>')
            notes_html = "".join(notes_parts)

            rows.append(f"""<tr>
  <td class="num">{i}</td>
  <td><strong>{h(f.principle)}</strong></td>
  <td>{status_badge(f.status)}</td>
  <td>{severity_badge(f.severity)}</td>
  <td class="evidence">{h(f.evidence)}</td>
  <td class="rec">{h(f.recommendation)}{notes_html}</td>
</tr>""")

        pillar_html_sections.append(f"""
<section class="pillar-section" id="pillar-{pname.lower().replace(' ', '-').replace('&', '')}">
  <div class="pillar-header">
    <div class="pillar-title-group">
      <h2>{h(pname)}</h2>
      <span class="compliance-tag" style="background:{color}15;color:{color};border:1px solid {color}30">{comp}</span>
    </div>
    <div class="pillar-gauge">
      <svg viewBox="0 0 36 36" class="circular-chart">
        <path class="circle-bg" d="M18 2.0845 a 15.9155 15.9155 0 0 1 0 31.831 a 15.9155 15.9155 0 0 1 0 -31.831"/>
        <path class="circle" stroke="{color}" stroke-dasharray="{sc * 10}, 100" d="M18 2.0845 a 15.9155 15.9155 0 0 1 0 31.831 a 15.9155 15.9155 0 0 1 0 -31.831"/>
        <text x="18" y="20.35" class="percentage">{sc:.0f}</text>
      </svg>
      <span class="gauge-label">/10</span>
    </div>
  </div>
  <div class="table-wrap">
    <table>
      <thead><tr><th>#</th><th>Principle</th><th>Status</th><th>Severity</th><th>Evidence</th><th>Recommendation</th></tr></thead>
      <tbody>{"".join(rows)}</tbody>
    </table>
  </div>
</section>""")

    # Priority actions
    action_items = []
    for p in pillars.values():
        for f in p.findings:
            if f.status in (Status.FAIL, Status.PARTIAL):
                action_items.append(f)
    severity_order = {Severity.CRITICAL: 0, Severity.HIGH: 1, Severity.MEDIUM: 2, Severity.LOW: 3, Severity.INFO: 4}
    action_items.sort(key=lambda f: (severity_order.get(f.severity, 9), 0 if f.status == Status.FAIL else 1))

    action_html_items = []
    for i, f in enumerate(action_items, 1):
        icon = "🔴" if f.status == Status.FAIL else "🟡"
        notes = ""
        if f.acceptable_risk_note:
            notes += f'<div class="note note-risk"><strong>⚡ Acceptable Risk:</strong> {h(f.acceptable_risk_note)}</div>'
        if f.alternative_note:
            notes += f'<div class="note note-alt"><strong>↔ Alternative:</strong> {h(f.alternative_note)}</div>'
        action_html_items.append(f"""
<div class="action-item">
  <div class="action-num">{i}</div>
  <div class="action-body">
    <div class="action-title">{icon} {severity_badge(f.severity)} <strong>{h(f.principle)}</strong> <span class="action-pillar">({h(f.pillar)})</span></div>
    <p>{h(f.recommendation)}</p>
    {notes}
  </div>
</div>""")

    # Build pillar nav
    nav_items = []
    for pname, pillar in pillars.items():
        pid = pname.lower().replace(' ', '-').replace('&', '')
        color = score_color(pillar.score)
        nav_items.append(
            f'<a href="#pillar-{pid}" class="nav-pill" style="border-color:{color}">'
            f'<span class="nav-score" style="color:{color}">{pillar.score:.0f}</span>'
            f'<span class="nav-name">{h(pname)}</span></a>'
        )

    return f"""<!DOCTYPE html>
<html lang="en" data-theme="light">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Well-Architected Report — {h(repo_name)}</title>
<style>
:root {{
  --bg: #ffffff; --bg2: #f8fafc; --bg3: #f1f5f9;
  --fg: #0f172a; --fg2: #475569; --fg3: #94a3b8;
  --border: #e2e8f0; --shadow: 0 1px 3px rgba(0,0,0,.08);
  --radius: 12px; --radius-sm: 8px;
  --accent: #6366f1;
}}
[data-theme="dark"] {{
  --bg: #0f172a; --bg2: #1e293b; --bg3: #334155;
  --fg: #f1f5f9; --fg2: #94a3b8; --fg3: #64748b;
  --border: #334155; --shadow: 0 1px 3px rgba(0,0,0,.3);
}}
*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', system-ui, sans-serif;
  background: var(--bg); color: var(--fg); line-height: 1.6;
  -webkit-font-smoothing: antialiased;
}}
.container {{ max-width: 1200px; margin: 0 auto; padding: 2rem 1.5rem; }}

/* Header */
.header {{
  background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 50%, #a78bfa 100%);
  padding: 3rem 0; color: white; position: relative; overflow: hidden;
}}
.header::before {{
  content: ''; position: absolute; top: -50%; right: -10%; width: 500px; height: 500px;
  background: rgba(255,255,255,.06); border-radius: 50%;
}}
.header::after {{
  content: ''; position: absolute; bottom: -30%; left: -5%; width: 300px; height: 300px;
  background: rgba(255,255,255,.04); border-radius: 50%;
}}
.header .container {{ position: relative; z-index: 1; }}
.header h1 {{ font-size: 2rem; font-weight: 700; margin-bottom: .25rem; letter-spacing: -.02em; }}
.header .subtitle {{ opacity: .85; font-size: .95rem; }}
.header-meta {{ display: flex; gap: 2rem; margin-top: 1rem; font-size: .85rem; opacity: .8; flex-wrap: wrap; }}

/* Theme toggle */
.theme-toggle {{
  position: fixed; top: 1rem; right: 1rem; z-index: 100;
  background: var(--bg2); border: 1px solid var(--border); border-radius: 50%;
  width: 40px; height: 40px; cursor: pointer; display: flex; align-items: center;
  justify-content: center; font-size: 1.2rem; box-shadow: var(--shadow);
  transition: all .2s;
}}
.theme-toggle:hover {{ transform: scale(1.1); }}

/* Score overview */
.score-overview {{
  display: grid; grid-template-columns: auto 1fr; gap: 2.5rem;
  background: var(--bg2); border: 1px solid var(--border); border-radius: var(--radius);
  padding: 2rem; margin: 2rem 0; box-shadow: var(--shadow); align-items: center;
}}
.main-score {{ text-align: center; }}
.main-score .circular-chart {{ width: 140px; height: 140px; }}
.main-score .score-num {{ font-size: 1.5rem; font-weight: 700; color: var(--fg); margin-top: .5rem; }}
.main-score .score-label {{ font-size: .85rem; color: var(--fg2); }}
.stat-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 1rem; }}
.stat-card {{
  background: var(--bg); border: 1px solid var(--border); border-radius: var(--radius-sm);
  padding: 1rem 1.25rem; text-align: center;
}}
.stat-card .stat-num {{ font-size: 1.75rem; font-weight: 700; }}
.stat-card .stat-label {{ font-size: .8rem; color: var(--fg2); margin-top: .25rem; }}

/* Navigation */
.pillar-nav {{
  display: flex; gap: .75rem; margin: 2rem 0; overflow-x: auto; padding-bottom: .5rem;
  flex-wrap: wrap;
}}
.nav-pill {{
  display: flex; align-items: center; gap: .5rem; padding: .6rem 1rem;
  background: var(--bg2); border: 2px solid; border-radius: 999px;
  text-decoration: none; color: var(--fg); font-size: .85rem; font-weight: 500;
  transition: all .2s; white-space: nowrap;
}}
.nav-pill:hover {{ transform: translateY(-2px); box-shadow: var(--shadow); }}
.nav-score {{ font-weight: 700; font-size: 1rem; }}

/* Pillar sections */
.pillar-section {{
  background: var(--bg2); border: 1px solid var(--border); border-radius: var(--radius);
  padding: 1.75rem; margin-bottom: 1.5rem; box-shadow: var(--shadow);
}}
.pillar-header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 1.25rem; }}
.pillar-title-group {{ display: flex; align-items: center; gap: .75rem; }}
.pillar-title-group h2 {{ font-size: 1.25rem; font-weight: 700; }}
.compliance-tag {{
  padding: .25rem .75rem; border-radius: 999px; font-size: .75rem;
  font-weight: 600; text-transform: uppercase; letter-spacing: .05em;
}}
.pillar-gauge {{ display: flex; align-items: center; gap: .25rem; }}
.pillar-gauge .circular-chart {{ width: 52px; height: 52px; }}
.gauge-label {{ font-size: .85rem; color: var(--fg2); font-weight: 600; }}

/* SVG gauge */
.circular-chart {{ display: block; }}
.circle-bg {{ fill: none; stroke: var(--border); stroke-width: 3; }}
.circle {{
  fill: none; stroke-width: 3; stroke-linecap: round;
  animation: progress 1s ease-out forwards;
  transform: rotate(-90deg); transform-origin: 50% 50%;
}}
@keyframes progress {{ 0% {{ stroke-dasharray: 0 100; }} }}
.percentage {{ fill: var(--fg); font-size: .55em; text-anchor: middle; font-weight: 700; }}

/* Tables */
.table-wrap {{ overflow-x: auto; }}
table {{ width: 100%; border-collapse: collapse; font-size: .85rem; }}
th {{ background: var(--bg3); padding: .7rem .75rem; text-align: left; font-weight: 600;
     font-size: .75rem; text-transform: uppercase; letter-spacing: .05em; color: var(--fg2); }}
td {{ padding: .7rem .75rem; border-top: 1px solid var(--border); vertical-align: top; }}
tr:hover td {{ background: var(--bg3); }}
.num {{ width: 2rem; text-align: center; color: var(--fg3); font-weight: 600; }}
.evidence {{ max-width: 250px; font-size: .8rem; color: var(--fg2); }}
.rec {{ max-width: 300px; font-size: .8rem; }}

/* Badges */
.badge {{
  display: inline-block; padding: .15rem .55rem; border-radius: 999px;
  font-size: .72rem; font-weight: 600; white-space: nowrap;
}}

/* Notes */
.note {{
  margin-top: .5rem; padding: .5rem .75rem; border-radius: var(--radius-sm);
  font-size: .78rem; line-height: 1.5;
}}
.note-risk {{ background: #f5f3ff; color: #6d28d9; border: 1px solid #ddd6fe; }}
.note-alt {{ background: #eff6ff; color: #1d4ed8; border: 1px solid #bfdbfe; }}
[data-theme="dark"] .note-risk {{ background: #2e1065; color: #c4b5fd; border-color: #4c1d95; }}
[data-theme="dark"] .note-alt {{ background: #172554; color: #93c5fd; border-color: #1e3a5f; }}

/* Action items */
.actions-section {{ margin: 2rem 0; }}
.actions-section h2 {{ font-size: 1.3rem; font-weight: 700; margin-bottom: 1rem; }}
.action-item {{
  display: flex; gap: 1rem; padding: 1rem; border-radius: var(--radius-sm);
  border: 1px solid var(--border); background: var(--bg2); margin-bottom: .75rem;
  transition: all .2s;
}}
.action-item:hover {{ border-color: var(--accent); box-shadow: 0 0 0 1px var(--accent)20; }}
.action-num {{
  min-width: 2rem; height: 2rem; background: var(--bg3); border-radius: 50%;
  display: flex; align-items: center; justify-content: center;
  font-weight: 700; font-size: .85rem; color: var(--fg2); flex-shrink: 0;
}}
.action-body {{ flex: 1; }}
.action-title {{ font-size: .9rem; margin-bottom: .35rem; display: flex; align-items: center; gap: .5rem; flex-wrap: wrap; }}
.action-pillar {{ color: var(--fg3); font-weight: 400; font-size: .82rem; }}
.action-body p {{ font-size: .83rem; color: var(--fg2); line-height: 1.6; }}

/* Footer */
.footer {{ text-align: center; padding: 2rem 0; color: var(--fg3); font-size: .8rem; border-top: 1px solid var(--border); margin-top: 2rem; }}
.footer a {{ color: var(--accent); text-decoration: none; }}

/* Responsive */
@media (max-width: 768px) {{
  .score-overview {{ grid-template-columns: 1fr; text-align: center; }}
  .stat-grid {{ grid-template-columns: repeat(2, 1fr); }}
  .header h1 {{ font-size: 1.5rem; }}
}}
@media print {{
  .theme-toggle {{ display: none; }}
  .header {{ background: #6366f1 !important; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
}}
</style>
</head>
<body>
<button class="theme-toggle" onclick="toggleTheme()" aria-label="Toggle theme">🌓</button>

<div class="header">
  <div class="container">
    <h1>Well-Architected Compliance Report</h1>
    <div class="subtitle">GitHub Well-Architected Framework Assessment</div>
    <div class="header-meta">
      <span>📦 {h(repo_name)}</span>
      <span>📅 {now}</span>
      <span>🔗 <a href="{FRAMEWORK_URL}" style="color:white;text-decoration:underline">{FRAMEWORK_URL}</a></span>
    </div>
  </div>
</div>

<div class="container">

  <!-- Score overview -->
  <div class="score-overview">
    <div class="main-score">
      <svg viewBox="0 0 36 36" class="circular-chart">
        <path class="circle-bg" d="M18 2.0845 a 15.9155 15.9155 0 0 1 0 31.831 a 15.9155 15.9155 0 0 1 0 -31.831"/>
        <path class="circle" stroke="{score_color(total / 5)}" stroke-dasharray="{total_pct}, 100"
              d="M18 2.0845 a 15.9155 15.9155 0 0 1 0 31.831 a 15.9155 15.9155 0 0 1 0 -31.831"/>
        <text x="18" y="20.35" class="percentage">{total_pct}</text>
      </svg>
      <div class="score-num">{total:.1f} / 50</div>
      <div class="score-label">Overall Score</div>
    </div>
    <div class="stat-grid">
      <div class="stat-card">
        <div class="stat-num" style="color:#10b981">{pass_count}</div>
        <div class="stat-label">Passing</div>
      </div>
      <div class="stat-card">
        <div class="stat-num" style="color:#f59e0b">{partial_count}</div>
        <div class="stat-label">Partial</div>
      </div>
      <div class="stat-card">
        <div class="stat-num" style="color:#ef4444">{fail_count}</div>
        <div class="stat-label">Failing</div>
      </div>
      <div class="stat-card">
        <div class="stat-num" style="color:#6b7280">{na_count}</div>
        <div class="stat-label">N/A</div>
      </div>
    </div>
  </div>

  <!-- Pillar nav -->
  <nav class="pillar-nav">
    {"".join(nav_items)}
  </nav>

  <!-- Pillar details -->
  {"".join(pillar_html_sections)}

  <!-- Priority actions -->
  <section class="actions-section" id="priority-actions">
    <h2>🎯 Priority Actions</h2>
    {"".join(action_html_items) if action_html_items else '<p style="color:var(--fg2)">No priority actions — all checks pass! 🎉</p>'}
  </section>

</div>

<div class="footer">
  Generated by the <strong>GitHub Well-Architected Assessment Tool</strong><br>
  Framework: <a href="{FRAMEWORK_URL}">{FRAMEWORK_URL}</a>
</div>

<script>
function toggleTheme() {{
  const html = document.documentElement;
  const current = html.getAttribute('data-theme');
  html.setAttribute('data-theme', current === 'dark' ? 'light' : 'dark');
  localStorage.setItem('wa-theme', html.getAttribute('data-theme'));
}}
// Restore saved theme
(function() {{
  const saved = localStorage.getItem('wa-theme');
  if (saved) document.documentElement.setAttribute('data-theme', saved);
  else if (window.matchMedia('(prefers-color-scheme: dark)').matches)
    document.documentElement.setAttribute('data-theme', 'dark');
}})();

// Smooth scroll for nav pills
document.querySelectorAll('.nav-pill').forEach(a => {{
  a.addEventListener('click', e => {{
    e.preventDefault();
    document.querySelector(a.getAttribute('href')).scrollIntoView({{ behavior: 'smooth', block: 'start' }});
  }});
}});
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Excel report generator
# ---------------------------------------------------------------------------

def generate_excel_report(pillars: dict[str, PillarScore], repo_name: str, output_path: Path) -> None:
    """Generate an Excel workbook with assessment results.

    Uses openpyxl if available, otherwise falls back to a CSV bundle.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback: generate CSV files instead
        _generate_csv_fallback(pillars, repo_name, output_path)
        return

    wb = Workbook()
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    # ── Style definitions ──
    header_font = Font(name="Aptos", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")

    status_fills = {
        "PASS": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        "PARTIAL": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "FAIL": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "ACCEPTABLE RISK": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
        "ALTERNATIVE APPROACH": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "N/A": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
    }
    status_fonts = {
        "PASS": Font(name="Aptos", color="166534"),
        "PARTIAL": Font(name="Aptos", color="92400E"),
        "FAIL": Font(name="Aptos", color="991B1B"),
        "ACCEPTABLE RISK": Font(name="Aptos", color="5B21B6"),
        "ALTERNATIVE APPROACH": Font(name="Aptos", color="1E40AF"),
        "N/A": Font(name="Aptos", color="6B7280"),
    }

    def style_header_row(ws, col_count: int) -> None:
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_data_cell(ws, row: int, col: int, status_val: str = "") -> None:
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = wrap_align
        if status_val in status_fills:
            cell.fill = status_fills[status_val]
            cell.font = status_fonts.get(status_val, Font(name="Aptos"))

    # ── Sheet 1: Summary ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = "6366F1"

    # Title area
    ws_summary.merge_cells("A1:F1")
    title_cell = ws_summary["A1"]
    title_cell.value = f"GitHub Well-Architected Assessment — {repo_name}"
    title_cell.font = Font(name="Aptos", bold=True, size=16, color="6366F1")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_summary.row_dimensions[1].height = 35

    ws_summary["A2"] = f"Date: {now}"
    ws_summary["A2"].font = Font(name="Aptos", color="6B7280", size=10)
    ws_summary["A3"] = f"Framework: {FRAMEWORK_URL}"
    ws_summary["A3"].font = Font(name="Aptos", color="6B7280", size=10)

    # Pillar scores table
    row = 5
    headers = ["Pillar", "Score", "Out Of", "Percentage", "Compliance"]
    for c, hdr in enumerate(headers, 1):
        ws_summary.cell(row=row, column=c, value=hdr)
    style_header_row(ws_summary, len(headers))
    # Re-style row 5 as header
    for c in range(1, len(headers) + 1):
        cell = ws_summary.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    total_score = 0.0
    for pname, pillar in pillars.items():
        row += 1
        ws_summary.cell(row=row, column=1, value=pname).font = Font(name="Aptos", bold=True)
        ws_summary.cell(row=row, column=2, value=pillar.score)
        ws_summary.cell(row=row, column=3, value=10)
        ws_summary.cell(row=row, column=4, value=round(pillar.score / 10 * 100, 1))
        comp_cell = ws_summary.cell(row=row, column=5, value=pillar.compliance)
        for c in range(1, 6):
            style_data_cell(ws_summary, row, c)
        # Color the compliance cell
        comp_val = pillar.compliance
        if comp_val == "COMPLIANT":
            comp_cell.fill = status_fills["PASS"]
            comp_cell.font = status_fonts["PASS"]
        elif comp_val == "PARTIAL":
            comp_cell.fill = status_fills["PARTIAL"]
            comp_cell.font = status_fonts["PARTIAL"]
        else:
            comp_cell.fill = status_fills["FAIL"]
            comp_cell.font = status_fonts["FAIL"]
        total_score += pillar.score

    # Total row
    row += 1
    ws_summary.cell(row=row, column=1, value="OVERALL").font = Font(name="Aptos", bold=True, size=12)
    ws_summary.cell(row=row, column=2, value=total_score).font = Font(name="Aptos", bold=True, size=12)
    ws_summary.cell(row=row, column=3, value=50).font = Font(name="Aptos", bold=True, size=12)
    ws_summary.cell(row=row, column=4, value=round(total_score / 50 * 100, 1)).font = Font(name="Aptos", bold=True, size=12)
    for c in range(1, 6):
        cell = ws_summary.cell(row=row, column=c)
        cell.border = thin_border
        cell.fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")

    # Column widths
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 12
    ws_summary.column_dimensions["C"].width = 10
    ws_summary.column_dimensions["D"].width = 14
    ws_summary.column_dimensions["E"].width = 18

    # ── Sheet 2: All Findings ──
    ws_all = wb.create_sheet("All Findings")
    ws_all.sheet_properties.tabColor = "3B82F6"

    all_headers = ["#", "Pillar", "Principle", "Status", "Severity", "Evidence", "Recommendation", "Acceptable Risk Note", "Alternative Note"]
    for c, hdr in enumerate(all_headers, 1):
        ws_all.cell(row=1, column=c, value=hdr)
    style_header_row(ws_all, len(all_headers))

    row_num = 2
    for pname, pillar in pillars.items():
        for finding in pillar.findings:
            ws_all.cell(row=row_num, column=1, value=row_num - 1)
            ws_all.cell(row=row_num, column=2, value=pname)
            ws_all.cell(row=row_num, column=3, value=finding.principle)
            ws_all.cell(row=row_num, column=4, value=finding.status.value)
            ws_all.cell(row=row_num, column=5, value=finding.severity.value)
            ws_all.cell(row=row_num, column=6, value=finding.evidence)
            ws_all.cell(row=row_num, column=7, value=finding.recommendation)
            ws_all.cell(row=row_num, column=8, value=finding.acceptable_risk_note)
            ws_all.cell(row=row_num, column=9, value=finding.alternative_note)
            for c in range(1, len(all_headers) + 1):
                style_data_cell(ws_all, row_num, c, finding.status.value if c == 4 else "")
            ws_all.cell(row=row_num, column=1).alignment = center_align
            row_num += 1

    # Column widths
    col_widths = [5, 22, 40, 20, 12, 50, 55, 50, 50]
    for i, w in enumerate(col_widths, 1):
        ws_all.column_dimensions[get_column_letter(i)].width = w

    # Auto-filter
    ws_all.auto_filter.ref = f"A1:I{row_num - 1}"

    # ── Sheet 3: Priority Actions ──
    ws_actions = wb.create_sheet("Priority Actions")
    ws_actions.sheet_properties.tabColor = "EF4444"

    act_headers = ["Priority", "Severity", "Pillar", "Principle", "Status", "Recommendation"]
    for c, hdr in enumerate(act_headers, 1):
        ws_actions.cell(row=1, column=c, value=hdr)
    style_header_row(ws_actions, len(act_headers))

    action_items = []
    for p in pillars.values():
        for f in p.findings:
            if f.status in (Status.FAIL, Status.PARTIAL):
                action_items.append(f)
    severity_order = {Severity.CRITICAL: 0, Severity.HIGH: 1, Severity.MEDIUM: 2, Severity.LOW: 3, Severity.INFO: 4}
    action_items.sort(key=lambda f: (severity_order.get(f.severity, 9), 0 if f.status == Status.FAIL else 1))

    for i, f in enumerate(action_items, 1):
        r = i + 1
        ws_actions.cell(row=r, column=1, value=i).alignment = center_align
        ws_actions.cell(row=r, column=2, value=f.severity.value)
        ws_actions.cell(row=r, column=3, value=f.pillar)
        ws_actions.cell(row=r, column=4, value=f.principle)
        ws_actions.cell(row=r, column=5, value=f.status.value)
        ws_actions.cell(row=r, column=6, value=f.recommendation)
        for c in range(1, len(act_headers) + 1):
            style_data_cell(ws_actions, r, c, f.status.value if c == 5 else "")

    act_widths = [10, 12, 22, 40, 18, 60]
    for i, w in enumerate(act_widths, 1):
        ws_actions.column_dimensions[get_column_letter(i)].width = w

    if action_items:
        ws_actions.auto_filter.ref = f"A1:F{len(action_items) + 1}"

    # ── Sheet per pillar ──
    for pname, pillar in pillars.items():
        safe_name = pname[:28]  # Excel sheet name max = 31 chars
        ws_p = wb.create_sheet(safe_name)
        tab_colors = {
            "Security": "DC2626", "Reliability": "2563EB",
            "Operational Excellence": "D97706", "Performance Efficiency": "059669",
            "Collaboration": "7C3AED",
        }
        ws_p.sheet_properties.tabColor = tab_colors.get(pname, "6B7280")

        p_headers = ["#", "Principle", "Status", "Severity", "Evidence", "Recommendation", "Acceptable Risk", "Alternative"]
        for c, hdr in enumerate(p_headers, 1):
            ws_p.cell(row=1, column=c, value=hdr)
        style_header_row(ws_p, len(p_headers))

        for i, f in enumerate(pillar.findings, 1):
            r = i + 1
            ws_p.cell(row=r, column=1, value=i).alignment = center_align
            ws_p.cell(row=r, column=2, value=f.principle)
            ws_p.cell(row=r, column=3, value=f.status.value)
            ws_p.cell(row=r, column=4, value=f.severity.value)
            ws_p.cell(row=r, column=5, value=f.evidence)
            ws_p.cell(row=r, column=6, value=f.recommendation)
            ws_p.cell(row=r, column=7, value=f.acceptable_risk_note)
            ws_p.cell(row=r, column=8, value=f.alternative_note)
            for c in range(1, len(p_headers) + 1):
                style_data_cell(ws_p, r, c, f.status.value if c == 3 else "")

        p_widths = [5, 40, 20, 12, 50, 55, 45, 45]
        for ci, w in enumerate(p_widths, 1):
            ws_p.column_dimensions[get_column_letter(ci)].width = w

        # Score footer
        footer_row = len(pillar.findings) + 3
        ws_p.cell(row=footer_row, column=1, value="Score:").font = Font(name="Aptos", bold=True)
        ws_p.cell(row=footer_row, column=2, value=f"{pillar.score:.1f} / 10 — {pillar.compliance}").font = Font(name="Aptos", bold=True, color="6366F1")

    wb.save(str(output_path))


def _generate_csv_fallback(pillars: dict[str, PillarScore], repo_name: str, output_path: Path) -> None:
    """Fallback: generate a CSV file when openpyxl is not installed."""
    import csv

    csv_path = output_path.with_suffix(".csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["#", "Pillar", "Pillar Score", "Compliance", "Principle", "Status",
                         "Severity", "Evidence", "Recommendation", "Acceptable Risk", "Alternative"])
        idx = 1
        for pname, pillar in pillars.items():
            for finding in pillar.findings:
                writer.writerow([
                    idx, pname, pillar.score, pillar.compliance,
                    finding.principle, finding.status.value, finding.severity.value,
                    finding.evidence, finding.recommendation,
                    finding.acceptable_risk_note, finding.alternative_note,
                ])
                idx += 1
    warn(f"openpyxl not installed — wrote CSV instead: {csv_path}")
    info("Install openpyxl for proper Excel output: pip install openpyxl")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="GitHub Well-Architected Framework — Organizational Assessment",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python well_architected_assessment.py\n"
            "  python well_architected_assessment.py --repo-path /path/to/repo\n"
            "  python well_architected_assessment.py --output report.md\n"
            "  python well_architected_assessment.py --no-interactive\n"
            "  python well_architected_assessment.py --no-html --no-excel\n"
        ),
    )
    parser.add_argument(
        "--repo-path",
        type=str,
        default=".",
        help="Path to the repository to assess (default: current directory)",
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="Base name/path for output files (default: well-architected-report). Extensions are added automatically.",
    )
    parser.add_argument(
        "--no-interactive",
        action="store_true",
        help="Skip interactive organizational questions (automated checks only)",
    )
    parser.add_argument(
        "--no-html",
        action="store_true",
        help="Skip HTML report generation",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Skip Excel report generation",
    )
    args = parser.parse_args()

    repo_path = Path(args.repo_path).resolve()
    if not repo_path.is_dir():
        print(f"Error: {repo_path} is not a valid directory.", file=sys.stderr)
        sys.exit(1)

    # Verify it's a git repo
    if not (repo_path / ".git").exists():
        warn(f"{repo_path} does not appear to be a Git repository.")
        if not ask_yes_no("Continue anyway?", default=False):
            sys.exit(0)

    assessment = WellArchitectedAssessment(repo_path, skip_interactive=args.no_interactive)
    report = assessment.run()

    # Determine output base path
    if args.output:
        base = Path(args.output)
        # Strip any extension the user included so we can add our own
        if base.suffix in (".md", ".html", ".xlsx", ".csv"):
            base = base.with_suffix("")
    else:
        base = repo_path / "well-architected-report"

    # 1. Markdown report
    md_path = base.with_suffix(".md")
    md_path.write_text(report, encoding="utf-8")
    success(f"Markdown report: {md_path}")

    # 2. HTML report
    if not args.no_html:
        html_content = generate_html_report(assessment.pillars, repo_path.name)
        html_path = base.with_suffix(".html")
        html_path.write_text(html_content, encoding="utf-8")
        success(f"HTML dashboard:  {html_path}")

    # 3. Excel report
    if not args.no_excel:
        xlsx_path = base.with_suffix(".xlsx")
        generate_excel_report(assessment.pillars, repo_path.name, xlsx_path)
        success(f"Excel workbook:  {xlsx_path}")

    print()
    info("Next steps:")
    info("  1. Review the report and prioritize Critical/High items")
    info("  2. Address gaps starting from the Priority Actions list")
    info("  3. Re-run the assessment after making changes to track progress")
    if not args.no_html:
        info(f"  4. Open the dashboard in a browser: open {base.with_suffix('.html')}")


if __name__ == "__main__":
    main()
